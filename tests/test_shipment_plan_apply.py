import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Border, PatternFill, Side

from modules.logistics.shipment_plan_apply.column_utils import resolve_cell_value
from modules.logistics.shipment_plan_apply.diff import run_and_capture_diff
from modules.logistics.shipment_plan_apply.planner import build_plan, apply_plan
from modules.logistics.shipment_plan_apply.product_lookup import ProductLookupError, load_product_lookup
from modules.logistics.shipment_plan_apply.purchase_book import PurchaseBook
from modules.logistics.shipment_plan_apply.shipment_summary import ShipmentSummaryBook
from modules.logistics.shipment_plan_apply.shipment_templates import parse_shipment_plan

REAL_DATA_DIR = Path("/Users/jingyuhuang/Documents/Work/闰科/物流仓库/采购汇总+发货计划")


# ---------------------------------------------------------------------------
# column_utils.resolve_cell_value：预览用的"公式尽量算出结果"逻辑
# ---------------------------------------------------------------------------


def test_resolve_cell_value_evaluates_local_arithmetic_and_concat(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # A2=采购单号(带字母，模拟真实订单号) B2=型号 C2=箱数 D2=箱容 E2=CBM公式
    # F2=标签(自引用型号) G2=拼接(采购单号+型号，模拟真实的 "&" 那一列)
    ws.append(["采购单号", "型号", "箱数", "箱容", "CBM", "标签", "拼接"])
    ws.append(["GH-2501009", "TD-RZ-419", 3, 12, "=C2*D2", "=+B2", "=A2&B2"])
    wb.save(tmp_path / "f.xlsx")
    wb2 = openpyxl.load_workbook(tmp_path / "f.xlsx", data_only=False)
    ws2 = wb2.active

    assert resolve_cell_value(ws2, 2, 5) == 3 * 12  # CBM 列，纯数字运算
    assert resolve_cell_value(ws2, 2, 6) == "TD-RZ-419"  # 标签 = 自引用型号（字符串也要能算）
    # "&" 拼接：真实数据里两边都是带字母的编号（比如订单号、型号），不是纯数字，之前的实现
    # 会因为替换后的表达式里出现字母被当成"不安全"而算不出来，这里就是回归这个问题
    assert resolve_cell_value(ws2, 2, 7) == "GH-2501009TD-RZ-419"


def test_resolve_cell_value_gives_up_on_external_or_function_formulas(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["=VLOOKUP(A1,B:C,2,0)", "=SUM(A1:A5)"])
    wb.save(tmp_path / "f.xlsx")
    wb2 = openpyxl.load_workbook(tmp_path / "f.xlsx", data_only=False)
    ws2 = wb2.active

    assert resolve_cell_value(ws2, 2, 1) is None
    assert resolve_cell_value(ws2, 2, 2) is None


def test_resolve_cell_value_passes_through_plain_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([42, "文本", None])
    assert resolve_cell_value(ws, 1, 1) == 42
    assert resolve_cell_value(ws, 1, 2) == "文本"
    assert resolve_cell_value(ws, 1, 3) is None


# ---------------------------------------------------------------------------
# shipment_templates
# ---------------------------------------------------------------------------


def _write_walmart_plan(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["店铺", "RK-SKU", "GTIN", "WM-SKU", "Item name", "预计发货数量"])
    ws.append(["CK-沃尔玛", "TD-348", "gtin1", "WM-1", "Table Lamp", 21])
    ws.append([None, "TD-392", "gtin2", "WM-2", "Table Lamp", 0])  # 数量为 0，应该报错
    ws.append([None, "TD-521", "gtin3", "WM-3", "Table Lamp", 30])
    wb.save(path)


def test_parse_walmart_plan(tmp_path):
    path = tmp_path / "walmart.xlsx"
    _write_walmart_plan(path)
    plan = parse_shipment_plan(path, "Sheet")
    assert plan.template_type == "walmart"
    assert len(plan.lines) == 2  # 数量为 0 的那行被判定非法，不计入 lines
    assert len(plan.errors) == 1
    assert "正数" in plan.errors[0]
    assert plan.lines[0].zd == "CK-沃尔玛"
    assert plan.lines[0].sku_kind == "RK"
    assert plan.lines[0].sku == "TD-348"
    assert plan.lines[1].zd == "CK-沃尔玛"  # 店铺是靠"沿用上一个非空值"填下来的


def _write_amazon_plan(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["发货计划", None, 1])
    ws.append(["店铺", "SKU", "US", "CA"])
    ws.append(["cinkeda", "TD-CKD-206", 30, None])
    ws.append([None, "TD-CK-584", None, 21])
    ws.append([None, "TD-CK-57", 21, 21])  # 同一行两个目的地都有数量，要拆成两条
    wb.save(path)


def test_parse_amazon_plan_splits_multi_destination_rows(tmp_path):
    path = tmp_path / "amazon.xlsx"
    _write_amazon_plan(path)
    plan = parse_shipment_plan(path, "Sheet")
    assert plan.template_type == "amazon"
    assert len(plan.lines) == 4
    dest_by_sku = {(l.sku, l.destination_label): l.quantity for l in plan.lines}
    assert dest_by_sku[("TD-CKD-206", "US")] == 30
    assert dest_by_sku[("TD-CK-584", "CA")] == 21
    assert dest_by_sku[("TD-CK-57", "US")] == 21
    assert dest_by_sku[("TD-CK-57", "CA")] == 21
    assert all(l.sku_kind == "AMZ" for l in plan.lines)


def _write_overseas_plan(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["图片", "海外仓-SKU", "9.2发货", None])
    ws.append([None, None, "CA1", "WF-RX-MD"])
    ws.append([None, "TD-158", 30, None])
    ws.append([None, "TD-244", 30, 60])  # 两个目的仓都有数量
    wb.save(path)


def test_parse_overseas_plan(tmp_path):
    path = tmp_path / "overseas.xlsx"
    _write_overseas_plan(path)
    plan = parse_shipment_plan(path, "Sheet")
    assert plan.template_type == "overseas"
    assert len(plan.lines) == 3
    dest_by_sku = {(l.sku, l.destination_label): l.quantity for l in plan.lines}
    assert dest_by_sku[("TD-158", "CA1")] == 30
    assert dest_by_sku[("TD-244", "CA1")] == 30
    assert dest_by_sku[("TD-244", "WF-RX-MD")] == 60
    assert all(l.zd == l.destination_label for l in plan.lines)  # 海外仓：ZD 就是目的仓列头


# ---------------------------------------------------------------------------
# product_lookup
# ---------------------------------------------------------------------------


def _write_product_info(path: Path, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["AMZ-SKU", "RK-SKU", "同款"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)


def test_product_lookup_resolve(tmp_path):
    path = tmp_path / "product.xlsx"
    _write_product_info(
        path,
        [
            ("AMZ-1", "RK-1", "HH-1"),
            ("AMZ-2", "RK-2", "HH-2"),
        ],
    )
    lookup = load_product_lookup(path)
    assert lookup.resolve("AMZ", "AMZ-1") == "HH-1"
    assert lookup.resolve("RK", "RK-2") == "HH-2"
    assert lookup.resolve("RK", "不存在") is None


def test_product_lookup_raises_on_conflicting_mapping(tmp_path):
    path = tmp_path / "product.xlsx"
    _write_product_info(
        path,
        [
            ("AMZ-1", "RK-1", "HH-1"),
            ("AMZ-1", "RK-9", "HH-9"),  # 同一个 AMZ-SKU 映射到两个不同货号
        ],
    )
    with pytest.raises(ProductLookupError):
        load_product_lookup(path)


# ---------------------------------------------------------------------------
# purchase_book
# ---------------------------------------------------------------------------


def _write_purchase_book(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单号", "采购日期", "型号", "订单数量", "数量单位", dt.datetime(2024, 1, 1), dt.datetime(2024, 2, 1), "未出货数量"])
    ws.append([None, None, None, None, None, "出货时间", "出货时间", None])
    ws.append(["PO-EARLY", dt.datetime(2024, 1, 1), "M1", 100, "pcs", 20, None, "=D3-SUM(F3:G3)"])
    ws.append(["PO-LATE", dt.datetime(2024, 6, 1), "M1", 50, "pcs", None, None, "=D4-SUM(F4:G4)"])
    wb.save(path)
    return wb


def test_purchase_book_allocates_earliest_order_first(tmp_path):
    path = tmp_path / "purchase.xlsx"
    _write_purchase_book(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    book = PurchaseBook(wb.active)

    orders = book.by_model["M1"]
    assert [o.order_no for o in orders] == ["PO-EARLY", "PO-LATE"]
    assert orders[0].initial_remaining == 80
    assert orders[1].initial_remaining == 50

    outcome = book.allocate("M1", 100)
    assert outcome.shortfall == 0
    assert [(a.row.order_no, a.quantity) for a in outcome.allocations] == [
        ("PO-EARLY", 80),
        ("PO-LATE", 20),
    ]


def test_purchase_book_reports_shortfall(tmp_path):
    path = tmp_path / "purchase.xlsx"
    _write_purchase_book(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    book = PurchaseBook(wb.active)

    outcome = book.allocate("M1", 500)
    assert outcome.shortfall == 500 - 80 - 50


def test_purchase_book_insert_date_column_preserves_formula_and_unrelated_rows(tmp_path):
    path = tmp_path / "purchase.xlsx"
    _write_purchase_book(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    book = PurchaseBook(ws)

    # 插一个已有两列中间的日期
    mid_col = book.find_or_create_date_column(dt.date(2024, 1, 15))
    assert ws.cell(row=book.header_row, column=mid_col - 1).value == dt.datetime(2024, 1, 1)
    assert ws.cell(row=book.header_row, column=mid_col + 1).value == dt.datetime(2024, 2, 1)

    outcome = book.allocate("M1", 10)
    for a in outcome.allocations:
        book.write_allocation(a, mid_col)

    wb.save(path)
    wb2 = openpyxl.load_workbook(path, data_only=False)
    book2 = PurchaseBook(wb2.active)
    row2 = book2.by_model["M1"][0]
    assert row2.initial_remaining == 80 - 10  # 重新加载后公式算出来的未出货数量要正确


def test_purchase_book_insert_date_column_preserves_formatting(tmp_path):
    # 回归测试：跟 shipment_summary 那边同一类问题——insert_cols 新建出来的整列没有任何格式，
    # 列宽这种"整列"级别的设置也不会跟着 insert_cols 自动往右挪。
    path = tmp_path / "purchase.xlsx"
    _write_purchase_book(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active

    blue = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    thin = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    for r in (1, 2, 3, 4):
        for c in (6, 7):  # 两个已有日期列 F, G
            ws.cell(row=r, column=c).fill = blue
            ws.cell(row=r, column=c).border = thin
    ws.column_dimensions["F"].width = 9.5
    ws.column_dimensions["G"].width = 9.5
    wb.save(path)

    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2.active
    book = PurchaseBook(ws2)
    mid_col = book.find_or_create_date_column(dt.date(2024, 1, 15))  # 插在 F,G 之间
    wb2.save(path)

    wb3 = openpyxl.load_workbook(path, data_only=False)
    ws3 = wb3.active
    from openpyxl.utils import get_column_letter

    assert ws3.column_dimensions["F"].width == 9.5  # 左边不受影响的列，原样不动
    assert ws3.column_dimensions[get_column_letter(mid_col)].width == 9.5  # 新插入的列，抄了邻居的列宽
    assert ws3.column_dimensions[get_column_letter(mid_col + 1)].width == 9.5  # 原来的 G 右移，列宽跟过去
    assert ws3.cell(3, mid_col).fill.fgColor.rgb == "000000FF"
    assert ws3.cell(3, mid_col).border.top.style == "thin"


# ---------------------------------------------------------------------------
# shipment_summary
# ---------------------------------------------------------------------------


def _write_summary_book(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
        "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号", "标签",
    ]
    ws.append(headers)
    ws.append(["PO-1", "M1", 5, 3, 15, None, "待定", "未发货", None, None, None, None, None, None, None, 7, "=+A2"])
    ws.append(["PO-2", "M2", 2, 3, 6, None, "待定", "未发货", None, None, None, None, None, None, None, 8, "=+A3"])
    ws.append([None] * 4 + ["=SUBTOTAL(9,E2:E3)"] + [None] * 12)  # 模拟表底的合计行
    wb.save(path)
    return wb


def test_shipment_summary_insert_above_preserves_formatting(tmp_path):
    # 回归测试：真实文件是有格式的（字体、填充色、边框、行高），insert_rows 新建出来的空行
    # 默认没有任何格式，而且行高这种"整行"级别的设置不会跟着 insert_rows 自动往下挪——之前
    # 就是这两个问题导致用户反馈"写入了新表但排版都没了"。
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"]
    ws.append(headers)
    ws.append(["PO-1", "M1", 5, 3, 15, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    ws.append(["PO-2", "M2", 2, 3, 6, None, "待定", "未发货", None, None, None, None, None, None, None, None])

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).fill = yellow
        ws.cell(row=2, column=c).border = thin
    ws.row_dimensions[2].height = 30
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    book = ShipmentSummaryBook(ws2)
    changes = book.apply_shipment("PO-1", "M1", 5, "ZD1", dt.date(2026, 9, 1))
    wb2.save(path)

    wb3 = openpyxl.load_workbook(path)
    ws3 = wb3.active
    new_row, pending_row = changes[0].new_row, changes[0].pending_row
    assert ws3.cell(new_row, 1).fill.fgColor.rgb == "00FFFF00"
    assert ws3.cell(new_row, 1).border.top.style == "thin"
    assert ws3.row_dimensions[new_row].height == 30
    assert ws3.cell(pending_row, 1).fill.fgColor.rgb == "00FFFF00"
    assert ws3.cell(pending_row, 1).border.top.style == "thin"
    assert ws3.row_dimensions[pending_row].height == 30


def test_shipment_summary_insert_above_reindexes_formulas(tmp_path):
    path = tmp_path / "summary.xlsx"
    _write_summary_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    book = ShipmentSummaryBook(ws)

    changes = book.apply_shipment("PO-1", "M1", 5, "ZD1", dt.date(2026, 9, 1))
    assert len(changes) == 1
    change = changes[0]
    assert change.kind == "insert_above"
    assert change.new_row == 2
    assert change.pending_row == 3
    assert change.pending_remaining_after == 10

    wb.save(path)
    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2.active

    # 新行：数量/ZD/发货时间/状态是新值，标签公式正确指向自己这一行
    assert ws2.cell(row=2, column=5).value == 5  # 数量
    assert ws2.cell(row=2, column=6).value == "ZD1"
    assert ws2.cell(row=2, column=17).value == "=+A2"

    # 原待定行下移到第 3 行：数量扣减，标签公式跟着改成指向第 3 行（不是还停在 =+A2）
    assert ws2.cell(row=3, column=5).value == 10
    assert ws2.cell(row=3, column=7).value == "待定"
    assert ws2.cell(row=3, column=17).value == "=+A3"

    # 原来第 3 行（PO-2/M2）被顶到第 4 行，它自己的公式也要跟着改成指向第 4 行
    assert ws2.cell(row=4, column=1).value == "PO-2"
    assert ws2.cell(row=4, column=17).value == "=+A4"

    # 表底合计行被顶到第 5 行，区间引用整体后移一位——连起始边界 E2 也要变成 E3，因为原来
    # 第 2 行的数据本身也被这次插入顶到了第 3 行，range 得跟着挪，不能只有区间末尾变
    assert ws2.cell(row=5, column=5).value == "=SUBTOTAL(9,E3:E4)"


def test_shipment_summary_convert_in_place_when_exact_match(tmp_path):
    path = tmp_path / "summary.xlsx"
    _write_summary_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    book = ShipmentSummaryBook(ws)

    changes = book.apply_shipment("PO-2", "M2", 6, "ZD9", dt.date(2026, 9, 1))
    assert len(changes) == 1
    change = changes[0]
    assert change.kind == "convert_in_place"
    assert change.new_row is None
    assert ws.cell(row=3, column=5).value == 6
    assert ws.cell(row=3, column=7).value == dt.datetime(2026, 9, 1)
    assert ws.cell(row=3, column=8).value == "未发货"


def test_shipment_summary_blank_fields_actually_clear_stale_values(tmp_path):
    # 回归测试：真实数据里"待定"行经常已经带着"编号"/"备注"这些字段的历史值（比如"无库存9"）——
    # _blank_fields 之前用 ws.cell(row, col, value=None) 清空，这个写法在 openpyxl 里是个坑：
    # value 只有不是 None 才会真的赋值，传 None 等于没传，格子会原样留着旧值。这里让待定行带上
    # 非空的 编号/备注，转正/拆分之后必须变成空，不能把旧备注误当成新记录的状态。
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"]
    ws.append(headers)
    ws.append(["PO-1", "M1", 5, 3, 15, None, "待定", "未发货", "旧仓库", "旧FBA", "旧追踪", "无库存9", "旧货代", "旧出货单", None, 3])
    ws.append(["PO-2", "M2", 2, 3, 6, None, "待定", "未发货", "旧仓库2", None, None, "另一条旧备注", None, None, None, 8])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    book = ShipmentSummaryBook(ws2)

    # insert_above：新行是从待定行复制出来的，旧的 仓库/FBA/追踪/备注/货代/出货单/编号 都不该带过去
    changes = book.apply_shipment("PO-1", "M1", 5, "ZD1", dt.date(2026, 9, 1))
    new_row = changes[0].new_row
    for col in (9, 10, 11, 12, 13, 14, 16):  # 仓库/FBA ID/追踪编号/备注/货代/出货单号/编号
        assert ws2.cell(row=new_row, column=col).value is None, f"col {col} 应该清空"

    # convert_in_place：原地转正的那一行自己带的旧值也要被清掉
    changes2 = book.apply_shipment("PO-2", "M2", 6, "ZD2", dt.date(2026, 9, 1))
    pending_row = changes2[0].pending_row
    for col in (9, 10, 11, 12, 13, 14, 16):
        assert ws2.cell(row=pending_row, column=col).value is None, f"col {col} 应该清空"


def test_shipment_summary_missing_box_capacity_still_updates_boxes(tmp_path):
    # 回归测试：_set_explicit_fields/拆分剩余量那两处算出来的箱数在箱容缺失时会是 None——
    # 之前同样因为 ws.cell(..., value=None) 是 no-op，箱容缺失的行拆分/转正之后箱数格子会
    # 留着旧值，跟新的「数量」对不上。这里箱容留空，箱数原来的旧值应该被清掉，不能留着旧数字。
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"])
    ws.append(["PO-1", "M1", 999, None, 15, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    book = ShipmentSummaryBook(ws2)

    changes = book.apply_shipment("PO-1", "M1", 5, "ZD1", dt.date(2026, 9, 1))
    new_row, pending_row = changes[0].new_row, changes[0].pending_row
    # 箱容缺失，箱数算不出来，应该是 None，不能留着模板行的旧箱数 999
    assert ws2.cell(row=new_row, column=3).value is None
    assert ws2.cell(row=pending_row, column=3).value is None
    assert ws2.cell(row=pending_row, column=5).value == 10  # 数量本身照样正确扣减


def test_shipment_summary_pending_row_not_found_raises(tmp_path):
    path = tmp_path / "summary.xlsx"
    _write_summary_book(path)
    wb = openpyxl.load_workbook(path)
    book = ShipmentSummaryBook(wb.active)
    with pytest.raises(Exception):
        book.apply_shipment("PO-不存在", "M1", 1, "ZD", dt.date(2026, 9, 1))


def test_shipment_summary_skips_zero_quantity_sibling_row(tmp_path):
    # 回归测试：真实数据里出现过同一个采购单号+型号同时有两行待定——一行剩 36 个、一行是
    # 历史遗留的 0 个（比如对应的采购记录后来被清零了）。之前的实现只按采购单号+型号找，
    # 谁先出现在表里就抢到谁，曾经把 21 个写进了那条本该是 0 的待定行，还把它的箱数从 0
    # 改掉了；真正有货的那条反而没动。现在应该自动跳过 0 数量的行，从有货的那行扣。
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"])
    # 先放那条 0 数量的（模拟它在表里排在前面，最容易被"谁先找到用谁"的旧逻辑误伤）
    ws.append(["PO-DUP", "M1", 0, 3, 0, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    ws.append(["PO-DUP", "M1", 12, 3, 36, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    book = ShipmentSummaryBook(wb2.active)

    changes = book.apply_shipment("PO-DUP", "M1", 21, "ZD1", dt.date(2026, 9, 1))
    assert len(changes) == 1
    assert changes[0].kind == "insert_above"
    assert changes[0].pending_remaining_after == 15

    ws2 = wb2.active
    # 那条 0 数量的待定行必须原封不动，一个字段都不能被碰
    zero_row_values = [ws2.cell(row=2, column=c).value for c in range(1, 9)]
    assert zero_row_values == ["PO-DUP", "M1", 0, 3, 0, None, "待定", "未发货"]


def test_shipment_summary_drains_across_multiple_pending_rows(tmp_path):
    # 同一个采购单号+型号同时有好几行待定是正常状态（都是同一批还没决定去哪的库存），要发的
    # 数量比其中一行多的话，应该继续从下一行扣，不是报错——两行加起来（10+20=30）够 25 就行。
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"])
    ws.append(["PO-MULTI", "M1", 10, 1, 10, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    ws.append(["PO-MULTI", "M1", 20, 1, 20, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path)
    book = ShipmentSummaryBook(wb2.active)

    changes = book.apply_shipment("PO-MULTI", "M1", 25, "ZD1", dt.date(2026, 9, 1))
    assert len(changes) == 2
    assert changes[0].kind == "convert_in_place"  # 第一行 10 个正好扣完
    assert changes[0].quantity == 10
    assert changes[1].kind == "insert_above"  # 第二行扣 15，剩 5 还待定
    assert changes[1].quantity == 15
    assert changes[1].pending_remaining_after == 5

    assert book.total_pending_quantity("PO-MULTI", "M1") == 5


def test_shipment_summary_total_exceeding_all_pending_rows_raises(tmp_path):
    path = tmp_path / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
               "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"])
    ws.append(["PO-DUP", "M1", 0, 3, 0, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    ws.append(["PO-DUP", "M1", 12, 3, 36, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    wb.save(path)
    wb2 = openpyxl.load_workbook(path)
    book = ShipmentSummaryBook(wb2.active)

    # 两行加起来一共 36，要发 999 应该直接报错，不能瞎猜
    with pytest.raises(Exception):
        book.apply_shipment("PO-DUP", "M1", 999, "ZD1", dt.date(2026, 9, 1))


def test_shipment_summary_quantity_exceeding_pending_raises(tmp_path):
    path = tmp_path / "summary.xlsx"
    _write_summary_book(path)
    wb = openpyxl.load_workbook(path)
    book = ShipmentSummaryBook(wb.active)
    # PO-1/M1 待定数量是 15（见 _write_summary_book），要发 999 个应该直接报错，而不是被当成
    # "刚好发完"原地转正、把数字写错
    with pytest.raises(Exception):
        book.apply_shipment("PO-1", "M1", 999, "ZD1", dt.date(2026, 9, 1))


# ---------------------------------------------------------------------------
# planner + diff：串起来的小型集成测试（纯合成数据，不依赖真实文件）
# ---------------------------------------------------------------------------


def test_planner_and_diff_end_to_end(tmp_path):
    product_path = tmp_path / "product.xlsx"
    _write_product_info(product_path, [("AMZ-1", "RK-1", "M1")])
    lookup = load_product_lookup(product_path)

    purchase_path = tmp_path / "purchase.xlsx"
    _write_purchase_book(purchase_path)
    purchase_wb = openpyxl.load_workbook(purchase_path, data_only=False)
    purchase_book = PurchaseBook(purchase_wb.active)

    # 发货计划汇总表里的待定行要跟采购汇总表对得上号（同一个采购单号+型号），这里手写一份，
    # 不能直接借用 _write_summary_book 那份（用的是 PO-1/PO-2，跟这里的 PO-EARLY 对不上）
    summary_path = tmp_path / "summary.xlsx"
    summary_setup_wb = openpyxl.Workbook()
    summary_ws = summary_setup_wb.active
    headers = [
        "采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态",
        "仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号",
    ]
    summary_ws.append(headers)
    summary_ws.append(["PO-EARLY", "M1", 27, 3, 80, None, "待定", "未发货", None, None, None, None, None, None, None, None])
    summary_setup_wb.save(summary_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    summary_book = ShipmentSummaryBook(summary_wb.active)

    from modules.logistics.shipment_plan_apply.shipment_templates import PlanLine

    lines = [
        PlanLine(zd="ZD1", sku_kind="RK", sku="RK-1", quantity=30, destination_label="ZD1", source_row=2)
    ]
    plan = build_plan(lines, [], lookup, purchase_book, dt.date(2026, 9, 1))
    assert not plan.has_blocking_errors

    result = run_and_capture_diff(plan, purchase_book, summary_book)
    assert result.purchase.before_rows[0]["未出货数量"] == 80
    assert result.purchase.after_rows[0]["未出货数量"] == 50
    assert len(result.summary.before_rows) == 1
    assert len(result.summary.after_rows) == 2  # 部分发货：新行 + 剩余待定行


def test_planner_blocks_whole_batch_on_shortfall(tmp_path):
    product_path = tmp_path / "product.xlsx"
    _write_product_info(product_path, [("AMZ-1", "RK-1", "M1")])
    lookup = load_product_lookup(product_path)

    purchase_path = tmp_path / "purchase.xlsx"
    _write_purchase_book(purchase_path)
    purchase_wb = openpyxl.load_workbook(purchase_path, data_only=False)
    purchase_book = PurchaseBook(purchase_wb.active)

    from modules.logistics.shipment_plan_apply.shipment_templates import PlanLine

    lines = [
        PlanLine(zd="ZD1", sku_kind="RK", sku="RK-1", quantity=9999, destination_label="ZD1", source_row=2)
    ]
    plan = build_plan(lines, [], lookup, purchase_book, dt.date(2026, 9, 1))
    assert plan.has_blocking_errors
    with pytest.raises(ValueError):
        apply_plan(plan, purchase_book, None)


# ---------------------------------------------------------------------------
# 真实数据集成测试（跑起来比较慢，日常开发用 -m "not slow" 跳过）
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not REAL_DATA_DIR.exists(), reason="需要本机真实的物流仓库数据文件")
def test_real_walmart_plan_allocates_against_real_purchase_book(tmp_path):
    import shutil

    lookup = load_product_lookup(REAL_DATA_DIR / "在售产品信息总表(测试用).xlsx")

    plan_path = REAL_DATA_DIR / "CK-Walmart9.1发货计划.xlsx"
    parsed = parse_shipment_plan(plan_path, "9.1")
    assert parsed.template_type == "walmart"
    assert not parsed.errors

    purchase_tmp = tmp_path / "purchase.xlsx"
    shutil.copy(REAL_DATA_DIR / "采购订单汇总表(测试用).xlsx", purchase_tmp)
    purchase_wb = openpyxl.load_workbook(purchase_tmp, data_only=False)
    purchase_book = PurchaseBook(purchase_wb.active)

    plan = build_plan(parsed.lines, parsed.errors, lookup, purchase_book, dt.date(2026, 9, 1))
    # 这份真实数据里 TD-640 这条已知缺货，整批应该被挡住
    assert plan.has_blocking_errors
    assert any("TD-RZ-585" in e for item in plan.items for e in item.errors)
