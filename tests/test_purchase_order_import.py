import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from PySide6.QtCore import QSettings

from modules.logistics.purchase_order_import.order_file import OrderFileError, parse_order_file
from modules.logistics.purchase_order_import.planner import build_plan, apply_plan
from modules.logistics.purchase_order_import.supplier_codes import SupplierCodeStore


# ---------------------------------------------------------------------------
# order_file.parse_order_file
# ---------------------------------------------------------------------------


def _write_order_file(path: Path, order_no="SX-2609209", supplier="东莞市盛鑫灯饰有限公司", rows=None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同 (2)"
    ws.append(["香港闰科有限公司"])
    ws.append(["采购订单"])
    ws.append([None, None, None, None, f"订单编号：{order_no}   PO：PO#{order_no}"])
    ws.append([
        "采购商（甲方）：香港闰科有限公司", None, None, None, "地址：xxx", "联系方式：", "139", None,
        "采购日期：", dt.datetime(2026, 9, 2),
    ])
    ws.append([f"供应商（乙方）：{supplier}", None, None, None, "地址：xxx", "联系方式：", None, None, "交货日期：", "详见附件"])
    ws.append(["序号", "sku", "产品名称", "图片IMG", "描述", "包装方式及尺寸", "数量", "单价", "交货日期", "备注"])
    for i, (sku, name, qty, delivery) in enumerate(rows or [], start=1):
        ws.append([i, sku, name, None, None, None, qty, 10, delivery, None])
    wb.save(path)


def test_parse_order_file_extracts_header_fields_and_lines(tmp_path):
    path = tmp_path / "order.xlsx"
    _write_order_file(
        path,
        rows=[
            ("TD-CY-410", "雅筑黑色五金台灯", 150, dt.datetime(2026, 9, 29)),
            ("TD-LO-92", "白门框台灯2p", 180, dt.datetime(2026, 10, 7)),
        ],
    )
    order = parse_order_file(path)
    assert order.order_no == "SX-2609209"
    assert order.supplier_name == "东莞市盛鑫灯饰有限公司"
    assert order.purchase_date == dt.date(2026, 9, 2)
    assert not order.errors
    assert len(order.lines) == 2
    assert order.lines[0].model == "TD-CY-410"
    assert order.lines[0].quantity == 150
    assert order.lines[0].delivery_date == dt.date(2026, 9, 29)


def test_parse_order_file_skips_blank_sku_rows(tmp_path):
    path = tmp_path / "order.xlsx"
    _write_order_file(path, rows=[("TD-CY-410", "灯", 10, dt.datetime(2026, 9, 29))])
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([3, None, None, None, None, None, None, None, None, None])  # sku 空，应该跳过
    wb.save(path)

    order = parse_order_file(path)
    assert len(order.lines) == 1


def test_parse_order_file_missing_order_no_raises(tmp_path):
    path = tmp_path / "order.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["没有订单编号的表格"])
    ws.append(["序号", "sku", "产品名称", "图片IMG", "描述", "包装方式及尺寸", "数量", "单价", "交货日期", "备注"])
    ws.append([1, "TD-1", "灯", None, None, None, 10, 5, dt.datetime(2026, 1, 1), None])
    wb.save(path)

    with pytest.raises(OrderFileError):
        parse_order_file(path)


def test_parse_order_file_missing_supplier_leaves_none_and_notes(tmp_path):
    path = tmp_path / "order.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None, None, "订单编号：AB-001"])
    ws.append(["序号", "sku", "产品名称", "图片IMG", "描述", "包装方式及尺寸", "数量", "单价", "交货日期", "备注"])
    ws.append([1, "TD-1", "灯", None, None, None, 10, 5, dt.datetime(2026, 1, 1), None])
    wb.save(path)

    order = parse_order_file(path)
    assert order.order_no == "AB-001"
    assert order.supplier_name is None
    assert order.purchase_date is None
    assert any("供应商" in e for e in order.errors)
    assert any("采购日期" in e for e in order.errors)


# ---------------------------------------------------------------------------
# supplier_codes.SupplierCodeStore
# ---------------------------------------------------------------------------


@pytest.fixture()
def settings(tmp_path):
    path = str(tmp_path / "settings.ini")
    return QSettings(path, QSettings.Format.IniFormat)


def test_supplier_code_store_roundtrip(settings):
    store = SupplierCodeStore(settings)
    assert store.mapping() == {}
    assert store.resolve("东莞市盛鑫灯饰有限公司") is None

    store.set_mapping({"东莞市盛鑫灯饰有限公司": "SX", "GH工厂": "GH"})
    assert store.mapping() == {"东莞市盛鑫灯饰有限公司": "SX", "GH工厂": "GH"}
    assert store.resolve("东莞市盛鑫灯饰有限公司") == "SX"
    assert store.resolve(None) is None

    store.set_mapping({})
    assert store.mapping() == {}


# ---------------------------------------------------------------------------
# planner.build_plan / apply_plan
# ---------------------------------------------------------------------------


def _write_purchase_summary(path: Path) -> None:
    # PurchaseBook 要求表头里有「数量单位」和「未出货数量」，且两者之间至少留一列日期列
    # （见 purchase_book.py），所以测试表里也要把这几列凑齐，不能只写 build_plan 用得到的
    # 那几列；PurchaseBook 还假设表头下一行是"出货时间"这种子表头行，真正的数据从表头
    # 再往下第二行才开始（见 purchase_book.py 的 sub_header_row / _load_rows），测试表结构
    # 要跟真实表一致（第1行标题、第2行表头、第3行子表头、第4行起才是数据），不然
    # PurchaseBook 会读不到任何一行数据。
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购订单汇总"
    ws.append([None] * 12)  # 第1行：标题/合计行占位
    ws.append([
        "序号", "订单号", "采购日期", "交货日期", "供应商名称", "店铺", "型号", "产品名称",
        "订单数量", "数量单位", "出货日期占位", "未出货数量",
    ])
    ws.append([None] * 10 + ["出货时间", None])  # 第3行：子表头
    ws.append([
        "001", "GH-2501009", dt.datetime(2025, 1, 7), dt.datetime(2025, 3, 17), "GH", None,
        "TD-RZ-419", "简约花瓶灰色树脂台灯", 180, "pcs", None, "=I4",
    ])
    # 给型号格子上个底色，模拟真实表格逐行有格式——新增行应该跟着抄这个格式
    ws.cell(row=4, column=7).fill = PatternFill("solid", fgColor="FFFF00")
    wb.save(path)


def _write_shipment_summary(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发货计划"
    for _ in range(4):
        ws.append([None] * 27)
    ws.append([
        "采购单号", "型号", "标签", "产品名称", "箱数", "箱容", "数量", "长", "宽", "高", "毛重",
        "CBM", "总材重", "总实重", "仓库", "FBA ID", "追踪编号", "ZD", "编号", "备注", "交货时间",
        "发货时间", "工厂", "DP", "货代", "出货单号", "状态", "so",
    ])
    # 一条历史记录：GH 工厂 + TD-RZ-419 型号，带箱容/长宽高/毛重，给"同工厂同型号"匹配用
    ws.append([
        "GH-2501009", "TD-RZ-419", "=+B6", "简约花瓶灰色树脂台灯", 60, 3, "=E6*F6", 550, 340, 440,
        13.6, None, None, None, "US", "FBA1", "TRACK1", "CA1", None, None, dt.datetime(2025, 3, 17),
        dt.datetime(2026, 1, 7), "GH", None, "KQ", "SK1", "已发货", None,
    ])
    ws.cell(row=6, column=4).fill = PatternFill("solid", fgColor="FFFF00")  # 产品名称格子的底色
    wb.save(path)


@pytest.fixture()
def tables(tmp_path):
    purchase_path = tmp_path / "purchase.xlsx"
    summary_path = tmp_path / "summary.xlsx"
    _write_purchase_summary(purchase_path)
    _write_shipment_summary(summary_path)
    return purchase_path, summary_path


def test_build_plan_matches_history_and_computes_boxes(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="GH-2609001",
        supplier="广东GH工厂",
        rows=[("TD-RZ-419", "简约花瓶灰色树脂台灯", 90, dt.datetime(2026, 10, 1))],
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {"广东GH工厂": "GH"})

    assert not plan.skipped_orders
    assert not plan.skipped_files
    assert len(plan.items) == 1
    item = plan.items[0]
    assert item.supplier_code == "GH"
    assert item.box_capacity == 3
    assert item.length == 550 and item.width == 340 and item.height == 440
    assert item.gross_weight == 13.6
    assert item.boxes == 30  # 90 / 3，整除
    assert item.boxes_exact is True
    assert not item.notes


def test_build_plan_new_product_leaves_dims_blank_with_note(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="GH-2609002",
        supplier="广东GH工厂",
        rows=[("TD-BRAND-NEW", "全新品", 100, dt.datetime(2026, 10, 1))],
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {"广东GH工厂": "GH"})

    item = plan.items[0]
    assert item.box_capacity is None
    assert item.length is None and item.width is None and item.height is None
    assert item.gross_weight is None
    assert item.boxes is None
    assert any("没有历史记录" in n for n in item.notes)


def test_build_plan_missing_supplier_mapping_leaves_code_blank(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="ZZ-001",
        supplier="没配过映射的供应商",
        rows=[("TD-X", "灯", 10, dt.datetime(2026, 10, 1))],
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {})

    item = plan.items[0]
    assert item.supplier_code is None
    assert any("映射代码" in n for n in item.notes)


def test_build_plan_skips_already_imported_order(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    # 这个订单号在采购汇总表里已经存在（GH-2501009）
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="GH-2501009",
        supplier="广东GH工厂",
        rows=[("TD-RZ-419", "灯", 10, dt.datetime(2026, 10, 1))],
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {"广东GH工厂": "GH"})

    assert not plan.items
    assert len(plan.skipped_orders) == 1
    assert plan.skipped_orders[0].order_no == "GH-2501009"


def test_build_plan_indivisible_boxes_note(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="GH-2609003",
        supplier="广东GH工厂",
        rows=[("TD-RZ-419", "灯", 91, dt.datetime(2026, 10, 1))],  # 91 / 3 除不尽
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {"广东GH工厂": "GH"})

    item = plan.items[0]
    assert item.boxes_exact is False
    assert any("除不尽" in n for n in item.notes)


def test_apply_plan_appends_rows_to_both_sheets_with_seq_numbering(tmp_path, tables):
    purchase_path, summary_path = tables
    order_folder = tmp_path / "orders"
    order_folder.mkdir()
    _write_order_file(
        order_folder / "order1.xlsx",
        order_no="GH-2609001",
        supplier="广东GH工厂",
        rows=[
            ("TD-RZ-419", "简约花瓶灰色树脂台灯", 90, dt.datetime(2026, 10, 1)),
            ("TD-NEW", "新品", 50, dt.datetime(2026, 10, 5)),
        ],
    )

    purchase_wb = openpyxl.load_workbook(purchase_path)
    summary_wb = openpyxl.load_workbook(summary_path)
    plan = build_plan(order_folder, purchase_wb.active, summary_wb.active, {"广东GH工厂": "GH"})
    apply_plan(plan, purchase_wb.active, summary_wb.active)

    p_ws = purchase_wb.active
    # 已有数据在第 4 行（表头第 2 行、子表头第 3 行），新行接着追加在第 5、6 行
    assert p_ws.cell(row=5, column=1).value == "002"  # 序号：已有 001 -> 002
    assert p_ws.cell(row=5, column=2).value == "GH-2609001"
    assert p_ws.cell(row=5, column=5).value == "GH"
    assert p_ws.cell(row=5, column=7).value == "TD-RZ-419"
    assert p_ws.cell(row=5, column=9).value == 90
    assert p_ws.cell(row=5, column=10).value == "pcs"
    assert p_ws.cell(row=5, column=6).value is None  # 店铺列模板行本来就是空的，抄过来还是空
    assert p_ws.cell(row=6, column=1).value == "003"
    # 未出货数量是公式，整行复制时应该按"自引用这一行"重新指向新行号（原来 =I4，新行是 =I5）
    assert p_ws.cell(row=5, column=12).value == "=I5"
    # 型号格子的底色格式也应该跟着抄过来
    assert p_ws.cell(row=5, column=7).fill.fgColor.rgb == p_ws.cell(row=4, column=7).fill.fgColor.rgb
    # 出货批次列（K 列，在数量单位和未出货数量之间）是模板行自己的出货历史，不该抄过来
    assert p_ws.cell(row=5, column=11).value is None

    s_ws = summary_wb.active
    # 已有数据在第 6 行，新行接着追加在第 7、8 行
    assert s_ws.cell(row=7, column=1).value == "GH-2609001"
    assert s_ws.cell(row=7, column=2).value == "TD-RZ-419"
    assert s_ws.cell(row=7, column=5).value == 30  # 箱数 = 90/3
    assert s_ws.cell(row=7, column=6).value == 3  # 箱容复制自历史
    assert s_ws.cell(row=7, column=8).value == 550  # 长
    assert s_ws.cell(row=7, column=22).value == "待定"  # 发货时间
    assert s_ws.cell(row=7, column=23).value == "GH"  # 工厂
    assert s_ws.cell(row=7, column=27).value == "未发货"  # 状态
    assert s_ws.cell(row=7, column=18).value is None  # ZD：模板行是"已发货"的旧值，不该抄过来
    assert s_ws.cell(row=7, column=3).value == "=+B7"  # 标签公式，自引用部分改指向新行
    assert s_ws.cell(row=7, column=7).value == "=E7*F7"  # 数量公式同理
    assert s_ws.cell(row=7, column=15).value is None  # 仓库：模板行是旧发货记录，不该抄过来
    assert s_ws.cell(row=7, column=16).value is None  # FBA ID 同理
    # 产品名称格子的底色格式也应该跟着抄过来
    assert s_ws.cell(row=7, column=4).fill.fgColor.rgb == s_ws.cell(row=6, column=4).fill.fgColor.rgb

    # 第二个型号是全新品，没有历史记录，长宽高/箱容/箱数都留空；但公式列还是照样抄公式
    assert s_ws.cell(row=8, column=2).value == "TD-NEW"
    assert s_ws.cell(row=8, column=5).value is None
    assert s_ws.cell(row=8, column=6).value is None
    assert s_ws.cell(row=8, column=3).value == "=+B8"
