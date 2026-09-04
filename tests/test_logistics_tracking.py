"""modules/logistics/logistics_tracking 的测试。

跟 shipment_plan_apply 一样分层测：
- tracking_pipeline.py 是纯业务逻辑（不 import PySide6），用 openpyxl.Workbook() 现造最小化
  的测试表格，不联网（联网查询那部分用假的 get_last_routes_for_carrier 替换掉）。
- platforms/registry.py 的多账号 fallback 逻辑，用假的 client factory 测，不连真实平台。
- credential_store.py 用指向临时文件的 QSettings(IniFormat) 测，不碰用户本机真实的存储。
"""
from __future__ import annotations

import openpyxl
import pytest
from openpyxl.styles import Border, Font, PatternFill, Side
from PySide6.QtCore import QSettings

from modules.logistics.logistics_tracking import tracking_pipeline
from modules.logistics.logistics_tracking.credential_store import CredentialStore, StoredAccount
from modules.logistics.logistics_tracking.platforms.base import RouteResult
from modules.logistics.logistics_tracking.platforms.registry import (
    get_last_routes_for_carrier,
)
from modules.logistics.logistics_tracking.tracking_pipeline import (
    HeaderNotFoundError,
    TrackingSheet,
)

# ---- 测试用的最小化表格 ----

_HEADERS = ["DD", "物流商", "运单号（必填）", "货件状态"]


def _make_ws(rows: list[tuple], headers: list[str] = _HEADERS):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * len(headers))  # 第1行留空，跟真实表格一样表头在第2行
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    return wb, ws


# ---- 表头识别 ----


def test_find_header_row_and_columns():
    wb, ws = _make_ws([("001", "HDJ", "HDJ0001", None)])
    sheet = TrackingSheet(ws)
    assert sheet.header_row == 2
    assert ws.cell(row=sheet.header_row, column=sheet.carrier_col).value == "物流商"
    assert ws.cell(row=sheet.header_row, column=sheet.waybill_col).value == "运单号（必填）"


def test_waybill_header_alias_fallback():
    headers = ["DD", "物流商", "物流单号", "货件状态"]  # 老表用的是"物流单号"不是"运单号（必填）"
    wb, ws = _make_ws([("001", "HP", "HP0001", None)], headers=headers)
    sheet = TrackingSheet(ws)
    assert ws.cell(row=sheet.header_row, column=sheet.waybill_col).value == "物流单号"


def test_missing_required_header_raises():
    headers = ["DD", "物流商", "运单号（必填）"]  # 缺"货件状态"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * len(headers))
    ws.append(headers)
    with pytest.raises(HeaderNotFoundError):
        TrackingSheet(ws)


def test_missing_waybill_header_raises():
    headers = ["DD", "物流商", "货件状态"]  # 没有任何运单号候选表头
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * len(headers))
    ws.append(headers)
    with pytest.raises(HeaderNotFoundError):
        TrackingSheet(ws)


# ---- 扫描范围 ----


def test_scan_rows_skips_completed_and_empty():
    wb, ws = _make_ws([
        ("001", "HDJ", "HDJ0001", None),  # 空状态，在范围内
        ("002", "HDJ", "HDJ0002", "2.24更新：已完成"),  # 含"完成"，跳过
        ("003", "HDJ", "HDJ0003", "运输中"),  # 在范围内
        (None, None, None, None),  # 完全空行，跳过
        ("005", "ZB", None, "运输中"),  # 没有运单号，跳过
    ])
    sheet = TrackingSheet(ws)
    rows = sheet.scan_rows()
    assert [r.row for r in rows] == [3, 5]
    assert [r.waybill for r in rows] == ["HDJ0001", "HDJ0003"]


# ---- build_preview（联网部分用假函数替换）----


def test_build_preview_groups_by_carrier_and_computes_has_update(monkeypatch):
    wb, ws = _make_ws([
        ("001", "HDJ", "HDJ0001", "运输中"),
        ("002", "HDJ", "HDJ0002", "运输中"),
        ("003", "NK", "NK0001", "运输中"),
    ])
    sheet = TrackingSheet(ws)
    # 预先在"最后流水"列写一个旧值，模拟这不是第一次跑——HDJ0001 旧值跟这次查到的一样，
    # 应该判定"无更新"；HDJ0002 没有旧值，应该判定"有更新"。
    sheet.last_route_col = sheet.status_col + 1
    ws.cell(row=sheet.header_row, column=sheet.last_route_col, value="最后流水")
    ws.cell(row=3, column=sheet.last_route_col, value="已到港")

    calls = []

    def fake_lookup(carrier, accounts, waybill_numbers):
        calls.append((carrier, accounts, tuple(waybill_numbers)))
        if carrier == "HDJ":
            return {
                "HDJ0001": RouteResult(waybill="HDJ0001", found=True, last_route="已到港"),
                "HDJ0002": RouteResult(waybill="HDJ0002", found=True, last_route="已签收"),
            }
        return {"NK0001": RouteResult(waybill="NK0001", error="未找到该运单")}

    monkeypatch.setattr(tracking_pipeline, "get_last_routes_for_carrier", fake_lookup)

    progress_calls = []
    previews = sheet.build_preview(
        {"HDJ": [("u", "p")]},
        progress_callback=lambda done, total: progress_calls.append((done, total)),
    )

    today = tracking_pipeline._today_label()
    by_waybill = {p.waybill: p for p in previews}
    assert by_waybill["HDJ0001"].new_has_update == f"{today}已查，无更新"
    assert by_waybill["HDJ0002"].new_has_update == f"{today}有更新"
    assert by_waybill["NK0001"].new_last_route is None
    assert by_waybill["NK0001"].new_has_update == "未找到该运单"

    assert len(calls) == 2  # 两个物流商各查一次，互相独立
    assert progress_calls[0] == (0, 2)
    assert progress_calls[-1] == (2, 2)


# ---- 写回（含新列的格式复制）----


def test_apply_preview_appends_columns_with_copied_style():
    wb, ws = _make_ws([("001", "HDJ", "HDJ0001", "运输中")])
    header_fill = PatternFill("solid", fgColor="FFCC00")
    header_font = Font(bold=True)
    border = Border(bottom=Side(style="thin"))
    for c in range(1, 5):
        cell = ws.cell(row=2, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    ws.column_dimensions[ws.cell(row=2, column=4).column_letter].width = 22

    sheet = TrackingSheet(ws)
    assert sheet.last_route_col is None
    assert sheet.has_update_col is None

    from modules.logistics.logistics_tracking.tracking_pipeline import TrackingPreviewRow

    previews = [
        TrackingPreviewRow(
            row=3, carrier="HDJ", waybill="HDJ0001", status="运输中",
            old_last_route=None, old_has_update=None,
            new_last_route="已到港", new_has_update="有更新",
        )
    ]
    sheet.apply_preview(previews)

    assert sheet.last_route_col == 5
    assert sheet.has_update_col == 6
    assert ws.cell(row=2, column=5).value == "最后流水"
    assert ws.cell(row=2, column=6).value == "是否有更新"
    assert ws.cell(row=3, column=5).value == "已到港"
    assert ws.cell(row=3, column=6).value == "有更新"

    # 新插入的表头格子要照抄相邻列的样式，不是 openpyxl 默认的空白格式
    new_header_cell = ws.cell(row=2, column=5)
    assert new_header_cell.fill.fgColor.rgb == header_fill.fgColor.rgb
    assert new_header_cell.font.bold is True
    assert ws.column_dimensions["E"].width == 22

    # 再跑一次不应该重复追加列——已经存在的表头要直接复用
    sheet2 = TrackingSheet(ws)
    assert sheet2.last_route_col == 5
    assert sheet2.has_update_col == 6
    sheet2.apply_preview(previews)
    assert ws.max_column == 6


# ---- platforms/registry.py 的多账号 fallback ----


class _FakeClient:
    def __init__(self, known: dict[str, RouteResult]):
        self._known = known

    def get_last_routes(self, waybill_numbers):
        return {wb: self._known[wb] for wb in waybill_numbers if wb in self._known}


def test_lookup_multi_account_falls_back_to_next_account(monkeypatch):
    from modules.logistics.logistics_tracking.platforms import registry as registry_module

    accounts_seen = []

    def fake_factory_first(username, password):
        accounts_seen.append(username)
        return _FakeClient({"HDJ0001": RouteResult(waybill="HDJ0001", found=True, last_route="A")})

    def fake_factory_second(username, password):
        accounts_seen.append(username)
        return _FakeClient({"HDJ0002": RouteResult(waybill="HDJ0002", found=True, last_route="B")})

    factories = [fake_factory_first, fake_factory_second]

    def dispatch(username, password):
        return factories[len(accounts_seen)](username, password)

    result = registry_module._lookup_multi_account(
        dispatch, [("acc1", "p1"), ("acc2", "p2")], ["HDJ0001", "HDJ0002", "HDJ0003"]
    )

    assert accounts_seen == ["acc1", "acc2"]
    assert result["HDJ0001"].last_route == "A"
    assert result["HDJ0002"].last_route == "B"
    assert result["HDJ0003"].found is False
    assert "已尝试全部账号" in result["HDJ0003"].error


def test_get_last_routes_for_carrier_without_accounts_reports_reason():
    result = get_last_routes_for_carrier("HDJ", [], ["HDJ0001"])
    assert "还没在「账号管理」里配置" in result["HDJ0001"].error


def test_get_last_routes_for_carrier_unimplemented_platform():
    result = get_last_routes_for_carrier("ZB", [("u", "p")], ["ZB0001"])
    assert "尚未接入自动查询" in result["ZB0001"].error


def test_get_last_routes_for_carrier_unknown_code():
    result = get_last_routes_for_carrier("XYZ", [], ["W1"])
    assert "未知物流商代码" in result["W1"].error


# ---- credential_store.py ----


@pytest.fixture()
def settings(tmp_path):
    path = str(tmp_path / "settings.ini")
    return QSettings(path, QSettings.Format.IniFormat)


def test_credential_store_add_reorder_delete(settings):
    store = CredentialStore(settings)
    assert store.accounts_for("HDJ") == []

    store.set_accounts("HDJ", [StoredAccount("新账号", "111"), StoredAccount("老账号", "222")])
    accounts = store.accounts_for("HDJ")
    assert [a.username for a in accounts] == ["新账号", "老账号"]

    # 调整顺序：老账号提到前面
    accounts[0], accounts[1] = accounts[1], accounts[0]
    store.set_accounts("HDJ", accounts)
    assert [a.username for a in store.accounts_for("HDJ")] == ["老账号", "新账号"]

    store.set_accounts("HDJ", [])
    assert store.accounts_for("HDJ") == []


def test_credential_store_accounts_by_platform_only_includes_configured(settings):
    store = CredentialStore(settings)
    store.set_accounts("HDJ", [StoredAccount("u1", "p1")])
    store.set_accounts("NK", [StoredAccount("u2", "p2"), StoredAccount("u3", "p3")])

    result = store.accounts_by_platform()
    assert result == {"HDJ": [("u1", "p1")], "NK": [("u2", "p2"), ("u3", "p3")]}
    assert "ZB" not in result  # 没配过账号的平台不出现
