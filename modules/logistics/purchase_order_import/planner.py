"""把"文件夹里的一批采购订单文件"变成"采购汇总表/发货计划汇总表各要新增哪些行"——分两步：

- build_plan()：只读，扫描文件夹、按订单号去重、给每个型号算供应商代码/箱数/长宽高毛重，
  不碰 openpyxl 的写操作。任何字段缺信息（供应商没映射、型号没有同工厂历史记录、箱数除不尽）
  都不是硬错误——直接留空，把情况记进这一行的 notes 里，跟 build_plan 结果一起摆给人看，
  不阻断其它行/其它订单继续处理（这一点跟 shipment_plan_apply 的"整批要么全过要么全不写"
  不一样：那边改的是已有库存记录，错一条会让数字对不上，这边是从零新增一整行，新增本身
  没有"对不上"的风险，缺信息也不影响别的行，所以做成尽量填、填不出来就留空+提示）。
- apply_plan()：在 build_plan() 的基础上，真的往两个 workbook 对象里追加新行（只在内存，
  不存盘，调用方自己决定什么时候 wb.save()）。两张表都是"接到表格现有数据最后一行继续往下追加"，
  不需要像 shipment_summary.py 那样在表格中间 insert_rows——那是"从已有的一条待定库存里分一部分
  出去"，这里是"这个型号在表里还完全没出现过"，直接接在最后一行下面写，不会打乱任何公式的
  行号引用（采购汇总表里 SUBTOTAL 的区间本来就写死到很大的行号，发货计划汇总表本身没有这种
  跨行公式，都不受追加影响）。

  新行不是凭空写一堆空格子——先用 column_utils.copy_row() 把"现有最后一行"整行复制过去
  （格式、公式都跟着抄一份，公式里"引用自己这一行"的部分会自动指向新行，见 copy_row 的
  说明），保证新行看起来跟旁边的行一样、公式列（店铺/标签/DP/CBM/未出货数量等）不用业务
  人员自己再拖一遍。复制完之后，再把这一行"每条订单各不相同"的字段（型号/数量/供应商代码
  等业务值）和"这一行是全新库存、不该继承模板行历史状态"的字段（采购汇总表的出货批次数量列、
  发货计划汇总表的仓库/FBA ID/追踪编号/ZD/编号/备注/货代/出货单号）覆盖成正确的值——前者
  覆盖成这一条记录自己的数据，后者统一覆盖成空，不然会把模板行"已经发生过的历史"错当成
  新订单的状态抄过来。

两张表要求的字段来源、留空的字段，见模块设计文档；这里只放实现。
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from ..shipment_plan_apply.column_utils import column_index_map, copy_row, reindex_formula, require_columns
from ..shipment_plan_apply.purchase_book import PurchaseBook
from ..shipment_plan_apply.shipment_summary import BLANK_FIELDS, NEW_STATUS, PENDING_LABEL, ShipmentSummaryBook
from .order_file import OrderFile, OrderFileError, list_order_files, parse_order_file

_UNIT = "pcs"
_SEQ_WIDTH = 3


@dataclass
class HistoricalDims:
    box_capacity: float | None
    length: float | None
    width: float | None
    height: float | None
    gross_weight: float | None
    source_row: int


@dataclass
class PlanItem:
    order_no: str
    model: str
    product_name: str
    quantity: int
    purchase_date: dt.date | None
    delivery_date: dt.date | None
    supplier_name: str | None
    supplier_code: str | None
    box_capacity: float | None
    length: float | None
    width: float | None
    height: float | None
    gross_weight: float | None
    boxes: float | None
    boxes_exact: bool
    source_file: str
    source_row: int
    notes: list[str] = field(default_factory=list)


@dataclass
class SkippedOrder:
    order_no: str
    source_file: str
    reason: str


@dataclass
class Plan:
    items: list[PlanItem]
    skipped_orders: list[SkippedOrder]
    skipped_files: list[str]  # 整份文件解析失败（比如提取不出订单号），文件名+原因已经拼进字符串里

    @property
    def total_lines(self) -> int:
        return len(self.items)


def _num(value) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return float(value)


def _compute_boxes(quantity: int, box_capacity: float | None) -> tuple[float | None, bool]:
    if not box_capacity or box_capacity <= 0:
        return None, False
    boxes = quantity / box_capacity
    exact = float(boxes).is_integer()
    return (int(boxes) if exact else boxes), exact


def _build_dims_index(ws: Worksheet, header_row: int, col: dict[str, int]) -> dict[tuple[str, str], HistoricalDims]:
    """按(工厂, 型号)建索引：同一个组合出现好几次的话，行号更大（更晚追加）的覆盖前面的，
    最终每个组合留下的是"最近一次"的箱容/长宽高/毛重。
    """
    c_model = col["型号"]
    c_factory = col["工厂"]
    c_capacity = col["箱容"]
    c_length = col["长"]
    c_width = col["宽"]
    c_height = col["高"]
    c_weight = col["毛重"]
    max_col = max(c_model, c_factory, c_capacity, c_length, c_width, c_height, c_weight)

    index: dict[tuple[str, str], HistoricalDims] = {}
    for row_no, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True), start=header_row + 1
    ):
        model = row[c_model - 1]
        factory = row[c_factory - 1]
        if not model or not factory:
            continue
        key = (str(factory).strip(), str(model).strip())
        index[key] = HistoricalDims(
            box_capacity=_num(row[c_capacity - 1]),
            length=_num(row[c_length - 1]),
            width=_num(row[c_width - 1]),
            height=_num(row[c_height - 1]),
            gross_weight=_num(row[c_weight - 1]),
            source_row=row_no,
        )
    return index


def _last_data_row(ws: Worksheet, header_row: int, key_col: int) -> int:
    """往下扫，找最后一行 key_col 有值的行号——表格实际数据结尾往往比 ws.max_row 小很多
    （尾部一大片只是带样式的空行），不能直接拿 ws.max_row 当追加位置。
    """
    last = header_row
    for row_no, value in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=key_col, values_only=True), start=header_row + 1
    ):
        if value[key_col - 1] is not None:
            last = row_no
    return last


def _next_seq(ws: Worksheet, header_row: int, seq_col: int, last_row: int) -> int:
    max_seq = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=seq_col, max_col=seq_col, values_only=True):
        try:
            max_seq = max(max_seq, int(str(row[0])))
        except (TypeError, ValueError):
            continue
    return max_seq + 1


def build_plan(
    folder: Path,
    purchase_ws: Worksheet,
    summary_ws: Worksheet,
    supplier_map: dict[str, str],
) -> Plan:
    purchase_book = PurchaseBook(purchase_ws)
    existing_order_nos = {r.order_no for r in purchase_book.rows}

    summary_book = ShipmentSummaryBook(summary_ws)
    summary_cols = column_index_map(summary_ws, summary_book.header_row)
    require_columns(summary_cols, ["型号", "工厂", "箱容", "长", "宽", "高", "毛重"], "发货计划汇总表")
    dims_index = _build_dims_index(summary_ws, summary_book.header_row, summary_cols)

    items: list[PlanItem] = []
    skipped_orders: list[SkippedOrder] = []
    skipped_files: list[str] = []
    seen_order_nos: set[str] = set()

    for path in list_order_files(folder):
        try:
            order = parse_order_file(path)
        except OrderFileError as exc:
            skipped_files.append(f"「{path.name}」：{exc}")
            continue

        if order.order_no in existing_order_nos or order.order_no in seen_order_nos:
            skipped_orders.append(
                SkippedOrder(order_no=order.order_no, source_file=path.name, reason="采购汇总表里已经存在这个订单号")
            )
            continue
        seen_order_nos.add(order.order_no)

        supplier_code = supplier_map.get(order.supplier_name.strip()) if order.supplier_name else None

        for line in order.lines:
            dims = dims_index.get((supplier_code, line.model)) if supplier_code else None
            box_capacity = dims.box_capacity if dims else None
            length = dims.length if dims else None
            width = dims.width if dims else None
            height = dims.height if dims else None
            gross_weight = dims.gross_weight if dims else None
            boxes, boxes_exact = _compute_boxes(line.quantity, box_capacity)

            item_notes: list[str] = []
            if order.supplier_name is None:
                item_notes.append("订单文件里没找到供应商信息，「供应商名称」「工厂」留空")
            elif supplier_code is None:
                item_notes.append(f"供应商「{order.supplier_name}」还没配置映射代码，「供应商名称」「工厂」留空")
            if supplier_code is not None and dims is None:
                item_notes.append(f"型号「{line.model}」在供应商「{supplier_code}」名下没有历史记录，箱容/长/宽/高/毛重留空")
            if order.purchase_date is None:
                item_notes.append("订单文件里没找到采购日期，留空")
            if boxes is not None and not boxes_exact:
                item_notes.append(f"订单数量 {line.quantity} 除以箱容 {box_capacity} 除不尽，箱数留了小数 {boxes}，需要人工核对")
            if line.delivery_date is None:
                item_notes.append("这一行没读到交货日期，留空")

            items.append(
                PlanItem(
                    order_no=order.order_no,
                    model=line.model,
                    product_name=line.product_name,
                    quantity=line.quantity,
                    purchase_date=order.purchase_date,
                    delivery_date=line.delivery_date,
                    supplier_name=order.supplier_name,
                    supplier_code=supplier_code,
                    box_capacity=box_capacity,
                    length=length,
                    width=width,
                    height=height,
                    gross_weight=gross_weight,
                    boxes=boxes,
                    boxes_exact=boxes_exact,
                    source_file=path.name,
                    source_row=line.source_row,
                    notes=item_notes,
                )
            )

    return Plan(items=items, skipped_orders=skipped_orders, skipped_files=skipped_files)


def _as_datetime(value: dt.date | None) -> dt.datetime | None:
    return dt.datetime.combine(value, dt.time()) if value is not None else None


def _apply_dim_column(
    ws: Worksheet, col: int, template_row: int, header_row: int, dest_row: int, fallback: float | None
) -> None:
    """长/宽/高这几列，发货计划汇总表里有的行是随箱型自动算的公式，有的是历史手填的写死
    数字——copy_row() 已经把模板行整行（含公式，自引用部分重指向新行）抄给新行了，模板行
    本身是公式的话这里什么都不用做，直接覆盖反而会把公式换成写死的历史数字，新行就不会再
    跟着箱型自动算了。只有模板行这一列不是公式时才需要额外处理：往上找最近一行这一列是
    公式的，把它的公式抄一份过去（同样重指向新行）；再往上也找不到公式，才退回到从同工厂
    同型号历史记录里查出来的写死数字。
    """
    template_value = ws.cell(row=template_row, column=col).value
    if isinstance(template_value, str) and template_value.startswith("="):
        return
    for r in range(template_row - 1, header_row, -1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.startswith("="):
            _set(ws, dest_row, col, reindex_formula(v, r, dest_row))
            return
    _set(ws, dest_row, col, fallback)


def _set(ws: Worksheet, row: int, col: int, value) -> None:
    # ws.cell(row, column, value=X) 在 openpyxl 里只有 X 不是 None 才会真的赋值——传 None
    # 等于没传，格子会保留 copy_row() 抄过来的旧值。这里统一走 .value = ，None 也能真的清空。
    ws.cell(row=row, column=col).value = value


def apply_plan(plan: Plan, purchase_ws: Worksheet, summary_ws: Worksheet) -> None:
    if not plan.items:
        return

    purchase_book = PurchaseBook(purchase_ws)
    p_cols = column_index_map(purchase_ws, purchase_book.header_row)
    require_columns(p_cols, ["序号", "订单号", "采购日期", "交货日期", "供应商名称", "型号", "产品名称", "订单数量", "数量单位"], "采购订单汇总表")
    p_last_row = max((r.row_index for r in purchase_book.rows), default=purchase_book.header_row)
    seq = _next_seq(purchase_ws, purchase_book.header_row, p_cols["序号"], p_last_row)

    # 同一个订单号在采购汇总表里只占一个序号——用 plan.items 里"同订单号的行连在一起"这个
    # 保证（build_plan 按订单一份一份处理，一个订单的所有型号行紧挨着追加），订单号变了才
    # 递增序号，行号range最后统一做单元格合并。
    item_seqs: list[int] = []
    current_seq = seq - 1
    prev_order_no: object = object()
    for item in plan.items:
        if item.order_no != prev_order_no:
            current_seq += 1
            prev_order_no = item.order_no
        item_seqs.append(current_seq)

    summary_book = ShipmentSummaryBook(summary_ws)
    s_cols = column_index_map(summary_ws, summary_book.header_row)
    require_columns(
        s_cols,
        ["采购单号", "型号", "产品名称", "箱数", "箱容", "长", "宽", "高", "毛重", "交货时间", "发货时间", "工厂", "状态"],
        "发货计划汇总表",
    )
    s_last_row = _last_data_row(summary_ws, summary_book.header_row, s_cols["采购单号"])

    p_template_row = p_last_row
    s_template_row = s_last_row

    # 采购汇总表「数量单位」和「未出货数量」之间那一大片是各批次的已出货数量——这些是模板行
    # 自己的出货历史，新订单还没发过货，这些格子照抄过来的话会凭空多出一堆假的已出货记录，
    # 必须显式清空（未出货数量本身是公式，留着不动，会在 Excel 里根据清空后的批次列自动算对）。
    p_batch_cols = range(purchase_book.date_col_start, purchase_book.date_col_end + 1)
    p_remark_col = p_cols.get("备注")

    for i, item in enumerate(plan.items):
        p_row = p_last_row + 1 + i
        copy_row(purchase_ws, dest_row=p_row, src_row=p_template_row)
        for c in p_batch_cols:
            _set(purchase_ws, p_row, c, None)
        if p_remark_col:
            _set(purchase_ws, p_row, p_remark_col, None)

        # 序号只写在同一订单号的第一行，后面几行留空，等下面统一合并单元格——跟纸质表格
        # 里"同一个订单号只写一次序号、其余行合并"的排版方式一致。
        is_first_of_order = i == 0 or plan.items[i - 1].order_no != item.order_no
        if is_first_of_order:
            _set(purchase_ws, p_row, p_cols["序号"], str(item_seqs[i]).zfill(_SEQ_WIDTH))
        else:
            _set(purchase_ws, p_row, p_cols["序号"], None)
        _set(purchase_ws, p_row, p_cols["订单号"], item.order_no)
        _set(purchase_ws, p_row, p_cols["采购日期"], _as_datetime(item.purchase_date))
        _set(purchase_ws, p_row, p_cols["交货日期"], _as_datetime(item.delivery_date))
        _set(purchase_ws, p_row, p_cols["供应商名称"], item.supplier_code)
        _set(purchase_ws, p_row, p_cols["型号"], item.model)
        _set(purchase_ws, p_row, p_cols["产品名称"], item.product_name)
        _set(purchase_ws, p_row, p_cols["订单数量"], item.quantity)
        _set(purchase_ws, p_row, p_cols["数量单位"], _UNIT)

        s_row = s_last_row + 1 + i
        copy_row(summary_ws, dest_row=s_row, src_row=s_template_row)
        # 仓库/FBA ID/追踪编号/编号/备注/货代/出货单号（BLANK_FIELDS）+ ZD：都是模板行"这一笔
        # 具体是怎么发出去的"记录，新订单还没分配到任何一次具体发货，照抄过来就是张冠李戴。
        for name in [*BLANK_FIELDS, "ZD"]:
            col = s_cols.get(name)
            if col:
                _set(summary_ws, s_row, col, None)

        _set(summary_ws, s_row, s_cols["采购单号"], item.order_no)
        _set(summary_ws, s_row, s_cols["型号"], item.model)
        _set(summary_ws, s_row, s_cols["产品名称"], item.product_name)
        _set(summary_ws, s_row, s_cols["箱数"], item.boxes)
        _set(summary_ws, s_row, s_cols["箱容"], item.box_capacity)
        _apply_dim_column(summary_ws, s_cols["长"], s_template_row, summary_book.header_row, s_row, item.length)
        _apply_dim_column(summary_ws, s_cols["宽"], s_template_row, summary_book.header_row, s_row, item.width)
        _apply_dim_column(summary_ws, s_cols["高"], s_template_row, summary_book.header_row, s_row, item.height)
        _set(summary_ws, s_row, s_cols["毛重"], item.gross_weight)
        _set(summary_ws, s_row, s_cols["交货时间"], _as_datetime(item.delivery_date))
        _set(summary_ws, s_row, s_cols["发货时间"], PENDING_LABEL)
        _set(summary_ws, s_row, s_cols["工厂"], item.supplier_code)
        _set(summary_ws, s_row, s_cols["状态"], NEW_STATUS)

    _merge_order_seq_cells(purchase_ws, plan.items, p_cols["序号"], p_last_row)


def _merge_order_seq_cells(ws: Worksheet, items: list[PlanItem], seq_col: int, first_new_row: int) -> None:
    """同一订单号新增的几行连在一起，把它们的「序号」列合并成一个单元格（值已经只写在
    第一行，其余行留空，交给合并单元格来显示成"同一个序号跨了好几行"）。
    """
    run_start = first_new_row + 1
    for i in range(1, len(items) + 1):
        row = first_new_row + i
        is_last = i == len(items) or items[i].order_no != items[i - 1].order_no
        if is_last:
            if row > run_start:
                ws.merge_cells(start_row=run_start, start_column=seq_col, end_row=row, end_column=seq_col)
            run_start = row + 1
