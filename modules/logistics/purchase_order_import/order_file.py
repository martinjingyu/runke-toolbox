"""解析一份采购订单文件（合同格式的 .xlsx）——供应商开的合同表格，不是我们自己定义的模板，
所以字段靠"扫描前几行找带关键字的单元格"这种方式提取，不是固定表头：

- 订单编号：跟"PO编号"写在同一个单元格里，形如"订单编号：SX-2609209   PO：PO#2609209"，
  正则从"订单编号"后面把编号截出来。
- 供应商全称：跟"（乙方）"写在同一个单元格里，形如"供应商（乙方）：东莞市盛鑫灯饰有限公司"。
- 采购日期：跟供应商/订单编号不一样，是"标签"和"值"分成两个相邻单元格（标签在某一格，
  值在它右边那一格），不是同一格里拼在一起。

这三样任何一样没找到都不算硬错误——订单号找不到的话这份文件没法用（整份跳过，见 planner.py），
供应商/采购日期找不到就留空，交给后面的分摊逻辑决定要不要提示。

订单里的每个型号是"合同"表格里的一行，表头行本身没有固定行号（前面有几行是订单/供应商信息），
用 sku/数量/交货日期 这三个表头去定位表头行在第几行。
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..shipment_plan_apply.column_utils import HeaderNotFoundError, column_index_map, find_header_row, require_columns

ITEM_HEADERS = ["sku", "产品名称", "数量", "交货日期"]

_ORDER_NO_RE = re.compile(r"订单编号[：:]\s*(\S+)")
_SUPPLIER_RE = re.compile(r"供应商[^：:]*[：:]\s*(.+)")
_LABEL_SCAN_ROWS = 10


@dataclass
class OrderLine:
    model: str
    product_name: str
    quantity: int
    delivery_date: dt.date | None
    source_row: int


@dataclass
class OrderFile:
    order_no: str
    supplier_name: str | None
    purchase_date: dt.date | None
    source_path: Path
    lines: list[OrderLine]
    errors: list[str] = field(default_factory=list)


class OrderFileError(Exception):
    """整份文件没法用（比如连订单编号都提取不出来），调用方应该跳过这份文件。"""


def _to_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _scan_label_rows(ws: Worksheet, max_row: int) -> list[list]:
    last_row = min(max_row, ws.max_row or 0)
    return [[c.value for c in row] for row in ws.iter_rows(min_row=1, max_row=last_row)]


def _extract_order_no(rows: list[list]) -> str | None:
    for row in rows:
        for value in row:
            if isinstance(value, str) and "订单编号" in value:
                m = _ORDER_NO_RE.search(value)
                if m:
                    return m.group(1).strip()
    return None


def _extract_supplier_name(rows: list[list]) -> str | None:
    for row in rows:
        for value in row:
            if isinstance(value, str) and "供应商" in value:
                m = _SUPPLIER_RE.search(value)
                if m:
                    name = m.group(1).strip()
                    if name:
                        return name
    return None


def _extract_purchase_date(rows: list[list]) -> dt.date | None:
    for row in rows:
        for col_idx, value in enumerate(row):
            if isinstance(value, str) and "采购日期" in value and col_idx + 1 < len(row):
                d = _to_date(row[col_idx + 1])
                if d is not None:
                    return d
    return None


def _parse_quantity(value) -> tuple[int, str | None]:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0, f"数量格式不对，读到的是「{value!r}」，不是数字"
    if value <= 0:
        return 0, f"数量必须是正数，读到的是 {value}"
    if not float(value).is_integer():
        return 0, f"数量不是整数：{value}"
    return int(value), None


def parse_order_file(path: Path) -> OrderFile:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        label_rows = _scan_label_rows(ws, _LABEL_SCAN_ROWS)
        order_no = _extract_order_no(label_rows)
        if order_no is None:
            raise OrderFileError(f"「{path.name}」里没找到「订单编号」，这份文件没法用")
        supplier_name = _extract_supplier_name(label_rows)
        purchase_date = _extract_purchase_date(label_rows)

        errors: list[str] = []
        if supplier_name is None:
            errors.append("没找到「供应商」信息，供应商名称留空")
        if purchase_date is None:
            errors.append("没找到「采购日期」，采购日期留空")

        try:
            header_row = find_header_row(ws, ITEM_HEADERS, max_scan_rows=_LABEL_SCAN_ROWS + 5)
        except HeaderNotFoundError as exc:
            raise OrderFileError(f"「{path.name}」找不到订单明细表头（{ITEM_HEADERS}）：{exc}") from exc

        cols = column_index_map(ws, header_row)
        idx = require_columns(cols, ITEM_HEADERS, f"「{path.name}」订单明细")
        c_sku, c_name, c_qty, c_date = idx["sku"], idx["产品名称"], idx["数量"], idx["交货日期"]

        lines: list[OrderLine] = []
        for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            sku_val = row[c_sku - 1].value
            if sku_val is None or not str(sku_val).strip():
                continue
            model = str(sku_val).strip()

            name_val = row[c_name - 1].value
            product_name = str(name_val).strip() if name_val is not None else ""

            qty, qty_err = _parse_quantity(row[c_qty - 1].value)
            if qty_err:
                errors.append(f"第{row_no}行（sku={model}）：{qty_err}，这一行跳过")
                continue

            delivery_date = _to_date(row[c_date - 1].value)
            if delivery_date is None:
                errors.append(f"第{row_no}行（sku={model}）：没读到「交货日期」，留空")

            lines.append(
                OrderLine(
                    model=model,
                    product_name=product_name,
                    quantity=qty,
                    delivery_date=delivery_date,
                    source_row=row_no,
                )
            )

        return OrderFile(
            order_no=order_no,
            supplier_name=supplier_name,
            purchase_date=purchase_date,
            source_path=path,
            lines=lines,
            errors=errors,
        )
    finally:
        wb.close()


def list_order_files(folder: Path) -> list[Path]:
    """列出文件夹下所有采购订单文件——排除 Excel 打开时生成的临时锁文件（~$ 开头）。"""
    return sorted(
        p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")
    )
