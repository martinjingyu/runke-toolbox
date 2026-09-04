"""这个工具涉及真实库存和采购数据，所有表头匹配都要求精确一致（不做模糊/关键词匹配）——
读不到预期的表头就直接报错，交给人去看是不是表格结构变了，而不是让软件自己猜一个位置。
"""
from __future__ import annotations

import re
from copy import copy as copy_style

from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

_CELL_REF_RE = re.compile(r"\b([A-Za-z]{1,3})(\d+)\b")
_FORMULA_REF_RE = re.compile(r"([A-Za-z]{1,3})(\d+)")
# 公式原文里，把所有"单元格引用"都挖掉之后，剩下的字符必须只有这些运算符/数字/空白——这样
# 才能保证公式本身就是"若干个单元格引用做加减乘除或者用 & 拼起来"，没有别的花活（函数调用、
# 跨表/外部工作簿引用、写死的文本常量）。这个检查要在"把单元格引用换成实际值"之前做：型号、
# 订单号这些实际值本身经常是带字母的字符串（比如 GH-2501009），如果检查放在替换之后做，
# 这些字母会被误判成"公式里出现了不该有的字符"，导致原本能算的公式也算不出来。
_SAFE_FORMULA_STRUCTURE_RE = re.compile(r"^[\d.\+\-\*/&\s]*$")


def resolve_cell_value(ws: Worksheet, row: int, col: int, _visiting: frozenset = frozenset()):
    """给预览用的：公式格子尽量算出真实结果，而不是把公式原文显示给人看。

    只处理"纯本行四则运算/字符串拼接"这种公式（比如 CBM/总材重/总实重/重量 这些都是拿同一行
    的长宽高箱数箱容算出来的，"标签"是 =+B405 这种直接引用同一行别的格子，"&"是 =A405&B405
    字符串拼接）——这些不需要真的跑一个 Excel 公式引擎，本质就是普通四则运算，可以直接算。

    真正跨表/跨工作簿的公式（VLOOKUP 查在售产品信息总表、外部链接的供应商对照表等）算不出来，
    直接返回 None，交给调用方显示成空——这类字段基本都是"型号不变、查出来的结果也不变"，
    显示成空之后正好会被"没有变化的列不显示"这条规则自动过滤掉，不会露出一堆算不出来的公式
    原文糊弄人。
    """
    key = (row, col)
    if key in _visiting:
        return None  # 循环引用，放弃

    cell = ws.cell(row=row, column=col)
    value = cell.value
    if not (isinstance(value, str) and value.startswith("=")):
        return value

    formula = value[1:]
    if "[" in formula or "!" in formula or "(" in formula:
        return None  # 外部工作簿引用 / 函数调用，算不了

    structure = _CELL_REF_RE.sub("", formula)
    if not _SAFE_FORMULA_STRUCTURE_RE.match(structure):
        return None  # 公式里有单元格引用/运算符之外的东西，不是能直接算的纯本行公式

    visiting = _visiting | {key}
    failed = False

    def repl(m: re.Match) -> str:
        nonlocal failed
        letters, digits = m.group(1), m.group(2)
        try:
            ref_col = column_index_from_string(letters.upper())
        except ValueError:
            failed = True
            return m.group(0)
        ref_row = int(digits)
        resolved = resolve_cell_value(ws, ref_row, ref_col, visiting)
        if resolved is None:
            failed = True
            return m.group(0)
        return repr(resolved)  # repr() 会正确转义里面的引号/特殊字符，替换进去 eval 是安全的

    substituted = _CELL_REF_RE.sub(repl, formula)
    if failed:
        return None
    substituted = substituted.strip()
    if substituted.startswith("+"):
        # Excel 里"=+B2"这种开头的加号只是习惯写法（尤其是从别的工作簿粘公式过来的），
        # 不代表真的要做加法——Python 对字符串做一元 + 会直接报错（int 不会，但字符串会），
        # 所以要先把这个没有意义的开头加号去掉，不然引用的格子只要不是数字就会算不出来
        substituted = substituted[1:]
    substituted = substituted.replace("&", "+")  # Excel 的字符串拼接符，换成 Python 的 +

    try:
        return eval(substituted, {"__builtins__": {}}, {})  # noqa: S307 - 公式结构已经白名单过滤过
    except Exception:
        return None


class HeaderNotFoundError(Exception):
    pass


def find_header_row(ws: Worksheet, required_headers: list[str], max_scan_rows: int = 10) -> int:
    """在前 max_scan_rows 行里找到同时包含所有 required_headers 的那一行，返回行号（1-indexed）。

    ws[row_idx] 这种按行号取值的写法，如果 row_idx 超过表格实际的行数（比如整张表只有 3 行，
    却要看第 5 行），在 read_only 模式下会直接抛 IndexError 而不是返回空行——所以扫描范围要用
    表格实际的行数封顶，不能不管三七二十一扫到 max_scan_rows。
    """
    last_row = min(max_scan_rows, ws.max_row or 0)
    for row_idx in range(1, last_row + 1):
        values = [cell.value for cell in ws[row_idx]]
        if all(h in values for h in required_headers):
            return row_idx
    raise HeaderNotFoundError(
        f"在前 {max_scan_rows} 行里没找到包含这些表头的行：{required_headers}"
    )


def column_index_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """表头文字 -> 列号（1-indexed）。同一个表头出现多次时，以第一次出现的为准。"""
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is not None and cell.value not in mapping:
            mapping[cell.value] = cell.column
    return mapping


def require_columns(mapping: dict[str, int], names: list[str], context: str) -> dict[str, int]:
    missing = [n for n in names if n not in mapping]
    if missing:
        raise HeaderNotFoundError(f"{context}缺少必须的表头：{missing}")
    return {n: mapping[n] for n in names}


def reindex_formula(formula: str, old_row: int, new_row: int) -> str:
    """公式原文里，把"引用自己这一行"的单元格引用（行号正好等于 old_row 的）换成 new_row，
    其它行号的引用原样保留——给"整行复制到别的行"用的，不是"插入一行、后面所有行整体下移"
    那种要把一大片引用都 +1 的场景（那种见 shipment_summary.py 的 _shift_formula_refs）。
    """

    def repl(m: re.Match) -> str:
        letters, digits = m.group(1), m.group(2)
        if int(digits) == old_row:
            return f"{letters}{new_row}"
        return m.group(0)

    return _FORMULA_REF_RE.sub(repl, formula)


def copy_row(ws: Worksheet, dest_row: int, src_row: int, max_col: int | None = None) -> None:
    """整行复制：值、公式（自引用部分按 src_row -> dest_row 重新指向，见 reindex_formula）、
    格式（字体/填充/边框/对齐/数字格式/保护）、行高都复制过去——用在"接在表格最后一行下面
    追加一条新记录，但要长得跟原来的行一样"这种场景，不用逐个字段判断"这一列是不是公式、
    要不要抄格式"。

    注意这只处理"整行复制"本身；复制完之后调用方通常还要在上面用具体的业务值覆盖某些列
    （比如型号、数量这些一行一个样的字段），这个函数不负责判断哪些列该覆盖。
    """
    max_col = max_col or ws.max_column
    for c in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=c)
        dest_cell = ws.cell(row=dest_row, column=c)
        value = src_cell.value
        if isinstance(value, str) and value.startswith("="):
            value = reindex_formula(value, src_row, dest_row)
        dest_cell.value = value
        if src_cell.has_style:
            dest_cell.font = copy_style(src_cell.font)
            dest_cell.fill = copy_style(src_cell.fill)
            dest_cell.border = copy_style(src_cell.border)
            dest_cell.alignment = copy_style(src_cell.alignment)
            dest_cell.number_format = src_cell.number_format
            dest_cell.protection = copy_style(src_cell.protection)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height


def read_row(ws: Worksheet, row: int, header_row: int) -> dict:
    """把某一行按表头文字读成 {表头: 值} 的字典，给预览界面展示用——公式格子尽量换成算出来的
    结果（见 resolve_cell_value），不是原样把公式文字甩给人看。
    """
    mapping = column_index_map(ws, header_row)
    return {name: resolve_cell_value(ws, row, col) for name, col in mapping.items()}
