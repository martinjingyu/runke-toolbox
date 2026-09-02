"""发货计划汇总表：每个"采购单号+型号"在这张表里，发货时间="待定"的那些行就是这笔订单还在
工厂/没安排走的库存余量——**同一个采购单号+型号完全可能同时有好几行待定**，这些行加起来的
数量总和应该跟采购汇总表里这笔订单的未出货数量对得上，行与行之间没有另外的区别，都是同一批
"还没决定发去哪"的库存。所以分摊的时候不是"找唯一一行改"，而是"按行顺序（先到先扣）从这些
待定行里依次扣，扣完一行不够再扣下一行"，跟 purchase_book.py 里从多笔采购订单里依次分摊是
一个道理。

单独扣一行的时候，只有两种结果：
1. 这次要扣的数量 < 这一行剩余数量 -> 在这一行正上方插入一条新行，新行记这次实际扣掉的量；
   这一行自己的数量减掉这个量（箱数跟着重算），还是待定，可能在后面的分摊里继续被扣。
2. 这次要扣的数量 = 这一行剩余数量（正好扣完）-> 不插入新行，直接把这一行的发货时间从"待定"
   改成这次的日期，其它字段按新行的规则原地更新。

如果这个采购单号+型号名下所有待定行加起来都不够这次要发的数量，说明两张表本身的数据就对不上
（比如采购汇总表显示还有货，发货计划汇总表这边却没记全），直接报错，不猜、不硬写。

新行的内容策略是"整行照抄待定行，再覆盖需要改的几个字段"——不是逐个列名去挑要不要抄。这是
因为实测发现这张表里很多列（标签、CBM、总材重、总实重、重量、DP、FN sku 等）本身是公式，
而且是"引用自己这一行"的公式（比如"标签"是 =+B405，"总材重"是按本行箱数/长宽高算出来的），
新行如果只是把这些公式的值原样抄一份文字过去，要么把公式变成写死的旧数字（CBM/总材重这些
本该随新行箱数自动变化的数就再也不会变了），要么公式里的行号还指着旧行（比如插到新行里的
"=+B405"，新行明明是别的行号，会指错地方）。所以这里统一处理：整行复制过去，遇到公式就把
公式里"跟着这一行走"的单元格引用（形如字母+旧行号）换成新行号，再原样保留公式本身——这样
CBM/总材重这些依赖箱数的字段会在 Excel 里用新行的箱数自动重新算出正确结果，不用我们自己猜
换算公式。
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet

from .column_utils import column_index_map, find_header_row, require_columns

# 明确要清空、留给后续人工填写的字段
BLANK_FIELDS = ["仓库", "FBA ID", "追踪编号", "备注", "货代", "出货单号", "so", "编号"]

REQUIRED_HEADERS = ["采购单号", "型号", "箱数", "箱容", "数量", "ZD", "发货时间", "状态", *BLANK_FIELDS]

PENDING_LABEL = "待定"
NEW_STATUS = "未发货"

_FORMULA_REF_RE = re.compile(r"([A-Za-z]{1,3})(\d+)")


class PendingRowNotFoundError(Exception):
    pass


class InconsistentQuantityError(Exception):
    pass


@dataclass
class ShipmentSummaryChange:
    kind: str  # "insert_above" 或 "convert_in_place"
    pending_row: int
    new_row: int | None  # kind=="insert_above" 时，新插入行的行号
    order_no: str
    model: str
    quantity: int
    box_capacity: float | None
    boxes: float | None
    boxes_exact: bool
    zd: str
    ship_date: dt.date
    pending_remaining_after: int | None  # kind=="insert_above" 时，待定行扣完之后剩多少


def _reindex_formula(formula: str, old_row: int, new_row: int) -> str:
    def repl(m: re.Match) -> str:
        letters, digits = m.group(1), m.group(2)
        if int(digits) == old_row:
            return f"{letters}{new_row}"
        return m.group(0)

    return _FORMULA_REF_RE.sub(repl, formula)


def _shift_formula_refs(formula: str, inserted_at: int) -> str:
    # 插入一行之后，公式里任何指向"插入点或插入点以下"的行引用都要 +1——不只是自己引用自己
    # 这种（比如"标签"=+B405），像表格最底下的合计行 =SUBTOTAL(9,E6:E27947) 这种跨很多行的
    # 区间引用，只要区间的边界落在插入点以下，也要跟着挪一位，不然合计会漏掉插入进去的那行。
    def repl(m: re.Match) -> str:
        letters, digits = m.group(1), m.group(2)
        row_num = int(digits)
        if row_num >= inserted_at:
            return f"{letters}{row_num + 1}"
        return m.group(0)

    return _FORMULA_REF_RE.sub(repl, formula)


class ShipmentSummaryBook:
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.header_row = find_header_row(ws, REQUIRED_HEADERS, max_scan_rows=10)
        cols = column_index_map(ws, self.header_row)
        self.col = require_columns(cols, REQUIRED_HEADERS, "发货计划汇总表")

    def pending_rows(self, order_no: str, model: str) -> list[int]:
        """按行顺序返回这个采购单号+型号名下所有"发货时间=待定"的行号（可能不止一个）。"""
        c_order = self.col["采购单号"]
        c_model = self.col["型号"]
        c_ship = self.col["发货时间"]
        return [
            r
            for r in range(self.header_row + 1, self.ws.max_row + 1)
            if (
                self.ws.cell(row=r, column=c_order).value == order_no
                and self.ws.cell(row=r, column=c_model).value == model
                and self.ws.cell(row=r, column=c_ship).value == PENDING_LABEL
            )
        ]

    def find_pending_row(self, order_no: str, model: str) -> int | None:
        """随便找一行待定行（不保证是哪一行，也不保证是唯一一行）——只用来做"这个采购单号+
        型号到底存不存在待定库存"这种存在性判断，分摊逻辑不应该用这个，见 apply_shipment。
        """
        rows = self.pending_rows(order_no, model)
        return rows[0] if rows else None

    def total_pending_quantity(self, order_no: str, model: str) -> int:
        return sum(self.read_quantity(r) for r in self.pending_rows(order_no, model))

    def read_quantity(self, row: int) -> int:
        value = self.ws.cell(row=row, column=self.col["数量"]).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return int(value)
        if isinstance(value, str) and value.startswith("="):
            # "数量"本身是公式（一般是 箱数*箱容），公式结果读不到，改成自己拿箱数*箱容算
            boxes = self.ws.cell(row=row, column=self.col["箱数"]).value
            capacity = self.ws.cell(row=row, column=self.col["箱容"]).value
            if isinstance(boxes, (int, float)) and isinstance(capacity, (int, float)):
                return int(boxes * capacity)
        raise ValueError(f"发货计划汇总表第 {row} 行的「数量」既不是数字也不是能识别的公式，读不出来")

    def apply_shipment(
        self, order_no: str, model: str, quantity: int, zd: str, ship_date: dt.date
    ) -> list[ShipmentSummaryChange]:
        """把 quantity 这么多货，从这个采购单号+型号名下的待定行里依次扣掉，可能一次扣完
        一行也可能要扣好几行（先到先扣），返回按处理顺序排列的改动列表。
        """
        total_available = self.total_pending_quantity(order_no, model)
        if not self.pending_rows(order_no, model):
            raise PendingRowNotFoundError(
                f"发货计划汇总表里找不到采购单号「{order_no}」+ 型号「{model}」发货时间=待定 的行"
            )
        if quantity > total_available:
            # 正常走到这里之前，采购汇总表那边已经确认这个采购单号+型号有足够的未出货数量，
            # 理论上不该出现要发的比这边所有待定行加起来还多——真出现了，说明两张表本身就对
            # 不上（比如发货计划汇总表这边库存没记全），不该硬着头皮把数字覆盖过去，宁可停
            # 下来报错让人看。
            raise InconsistentQuantityError(
                f"发货计划汇总表里，采购单号「{order_no}」+ 型号「{model}」名下所有待定行的数量"
                f"加起来只有 {total_available}，但要写入的数量是 {quantity}，比这还多，数据对不上，"
                f"不能写入"
            )

        changes: list[ShipmentSummaryChange] = []
        remaining_need = quantity
        while remaining_need > 0:
            # 每一轮都重新找"当前还有余量的第一行"，不能缓存行号列表——插入行会让后面的行号
            # 整体往下挪，缓存的行号会失效。
            row = next(
                (r for r in self.pending_rows(order_no, model) if self.read_quantity(r) > 0),
                None,
            )
            if row is None:
                # total_available 已经在前面校验过够，理论上不会走到这里；真出现了说明前面的
                # 校验和这里的实际扣减对不上，同样不该瞎猜，直接报错。
                raise InconsistentQuantityError(
                    f"发货计划汇总表里，采购单号「{order_no}」+ 型号「{model}」的待定行在处理过程中"
                    f"意外用完了，还差 {remaining_need} 没能分摊，数据可能有问题，需要人工核对"
                )
            row_qty = self.read_quantity(row)
            take = min(row_qty, remaining_need)
            changes.append(self._consume_row(row, take, order_no, model, zd, ship_date))
            remaining_need -= take

        return changes

    def _consume_row(
        self, pending_row: int, quantity: int, order_no: str, model: str, zd: str, ship_date: dt.date
    ) -> ShipmentSummaryChange:
        """从这一行待定库存里扣 quantity（调用方保证 0 < quantity <= 这一行当前数量）。"""
        pending_qty = self.read_quantity(pending_row)
        box_capacity = self.ws.cell(row=pending_row, column=self.col["箱容"]).value
        boxes, boxes_exact = _compute_boxes(quantity, box_capacity)

        if quantity < pending_qty:
            new_row = pending_row
            template_row = pending_row + 1  # 待定行内容因为插入整体下移了一行
            self.ws.insert_rows(new_row)
            # insert_rows 只搬单元格的值，不会像 Excel 那样把公式里"指向自己这一行"的引用
            # 跟着往下调整——插入点以下所有行（不只是待定行）都要先把这类公式修好，不然
            # 插入点以下所有行的 CBM/总材重/总实重/DP 等公式都会读到错位的旧数据。
            self._reindex_shifted_rows(new_row)
            self._copy_row_with_reindex(new_row, template_row)
            self._set_explicit_fields(new_row, order_no, model, quantity, boxes, zd, ship_date, NEW_STATUS)
            self._blank_fields(new_row)

            remaining = pending_qty - quantity
            remaining_boxes, _ = _compute_boxes(remaining, box_capacity)
            self.ws.cell(row=template_row, column=self.col["数量"], value=remaining)
            self.ws.cell(row=template_row, column=self.col["箱数"], value=remaining_boxes)

            return ShipmentSummaryChange(
                kind="insert_above",
                pending_row=template_row,
                new_row=new_row,
                order_no=order_no,
                model=model,
                quantity=quantity,
                box_capacity=box_capacity,
                boxes=boxes,
                boxes_exact=boxes_exact,
                zd=zd,
                ship_date=ship_date,
                pending_remaining_after=remaining,
            )

        # quantity == pending_qty：正好扣完这一行，原地转正
        self._set_explicit_fields(pending_row, order_no, model, quantity, boxes, zd, ship_date, NEW_STATUS)
        self._blank_fields(pending_row)
        return ShipmentSummaryChange(
            kind="convert_in_place",
            pending_row=pending_row,
            new_row=None,
            order_no=order_no,
            model=model,
            quantity=quantity,
            box_capacity=box_capacity,
            boxes=boxes,
            boxes_exact=boxes_exact,
            zd=zd,
            ship_date=ship_date,
            pending_remaining_after=None,
        )

    def _reindex_shifted_rows(self, inserted_at: int) -> None:
        # inserted_at 这一行是刚插入的空行；它往下的每一行内容都往下搬了一行。这张表里的公式
        # 既有"自己引用自己这一行"的（比如"标签"=+B405），也有跨很多行的区间引用（比如最底下
        # 的合计行 =SUBTOTAL(9,E6:E27947)）——只要公式里出现的行号 >= 插入点，就该 +1，
        # 不管是哪一种。
        max_col = self.ws.max_column
        max_row = self.ws.max_row
        for row in range(inserted_at + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = self.ws.cell(row=row, column=c)
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    new_value = _shift_formula_refs(value, inserted_at)
                    if new_value != value:
                        cell.value = new_value

    def _copy_row_with_reindex(self, dest_row: int, src_row: int) -> None:
        max_col = self.ws.max_column
        for c in range(1, max_col + 1):
            value = self.ws.cell(row=src_row, column=c).value
            if isinstance(value, str) and value.startswith("="):
                value = _reindex_formula(value, src_row, dest_row)
            self.ws.cell(row=dest_row, column=c, value=value)

    def _set_explicit_fields(
        self, row: int, order_no: str, model: str, quantity: int, boxes, zd: str, ship_date: dt.date, status: str
    ) -> None:
        self.ws.cell(row=row, column=self.col["采购单号"], value=order_no)
        self.ws.cell(row=row, column=self.col["型号"], value=model)
        self.ws.cell(row=row, column=self.col["数量"], value=quantity)
        self.ws.cell(row=row, column=self.col["箱数"], value=boxes)
        self.ws.cell(row=row, column=self.col["ZD"], value=zd)
        self.ws.cell(row=row, column=self.col["发货时间"], value=dt.datetime.combine(ship_date, dt.time()))
        self.ws.cell(row=row, column=self.col["状态"], value=status)

    def _blank_fields(self, row: int) -> None:
        for field in BLANK_FIELDS:
            col = self.col.get(field)
            if col is not None:
                self.ws.cell(row=row, column=col, value=None)


def _compute_boxes(quantity: int, box_capacity) -> tuple[float | None, bool]:
    if not isinstance(box_capacity, (int, float)) or box_capacity <= 0:
        return None, False
    boxes = quantity / box_capacity
    exact = float(boxes).is_integer()
    return (int(boxes) if exact else boxes), exact
