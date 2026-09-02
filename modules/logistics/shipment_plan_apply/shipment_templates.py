"""解析运营提供的发货计划表——沃尔玛/亚马逊/海外仓三种模板，长得完全不一样：

- 沃尔玛：一行一个 RK-SKU，一个"预计发货数量"列，"店铺"整份文件基本固定。
- 亚马逊：一行一个 SKU（这里填的其实是 AMZ-SKU），"店铺"只在第一行填、后面靠合并单元格
  沿用；SKU 后面跟着若干个目的地列（列头是国家代码，比如 US/CA），同一行可能好几个目的地
  都有数量——每个目的地当成一条独立记录处理（各自扣库存、各自建一条待发货记录）。
- 海外仓：一行一个"海外仓-SKU"，表头分两行：第一行是固定列名，第二行是各个目的仓的 ZD 编号；
  同一行可能好几个目的仓都有数量，处理方式跟亚马逊的多目的地一样，拆成独立记录。

模板类型和用哪个 sheet 都不在这里自动决定"确定生效"——探测函数给出的只是"建议值"，
真正采不采用由上层（面板 UI）在预览界面里明确交给人确认，这里只负责"解析"。
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .column_utils import HeaderNotFoundError, column_index_map, find_header_row, require_columns

TemplateType = Literal["walmart", "amazon", "overseas"]

WALMART_HEADERS = ["店铺", "RK-SKU", "预计发货数量"]
AMAZON_HEADERS = ["店铺", "SKU"]
OVERSEAS_HEADERS = ["海外仓-SKU"]


@dataclass
class PlanLine:
    zd: str
    sku_kind: str  # "AMZ" 或 "RK"
    sku: str
    quantity: int
    destination_label: str
    source_row: int
    source_file: str = ""  # 一次可以导入好几份发货计划表，这个字段用来在报错/预览里区分是哪一份


@dataclass
class ParsedPlan:
    template_type: TemplateType
    lines: list[PlanLine]
    errors: list[str]


def list_sheet_names(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    return wb.sheetnames


def detect_template_type(ws: Worksheet) -> TemplateType | None:
    # ws[row_idx] 超过表格实际行数会直接抛 IndexError（read_only 模式下），扫描范围要用表格
    # 实际的行数封顶，不能不管三七二十一扫到第 5 行
    last_row = min(5, ws.max_row or 0)
    for row_idx in range(1, last_row + 1):
        values = {c.value for c in ws[row_idx] if c.value is not None}
        if set(WALMART_HEADERS) <= values:
            return "walmart"
    for row_idx in range(1, last_row + 1):
        values = {c.value for c in ws[row_idx] if c.value is not None}
        if set(OVERSEAS_HEADERS) <= values:
            return "overseas"
    for row_idx in range(1, last_row + 1):
        values = {c.value for c in ws[row_idx] if c.value is not None}
        if set(AMAZON_HEADERS) <= values:
            return "amazon"
    return None


def parse_shipment_plan(
    path: Path, sheet_name: str, template_type: TemplateType | None = None
) -> ParsedPlan:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    resolved_type = template_type or detect_template_type(ws)
    if resolved_type is None:
        raise HeaderNotFoundError(
            "没能自动识别出这是沃尔玛/亚马逊/海外仓哪一种发货计划表模板，"
            "需要手动指定模板类型"
        )

    if resolved_type == "walmart":
        header_row = find_header_row(ws, WALMART_HEADERS)
        lines, errors = _parse_walmart(ws, header_row)
    elif resolved_type == "amazon":
        header_row = find_header_row(ws, AMAZON_HEADERS)
        lines, errors = _parse_amazon(ws, header_row)
    elif resolved_type == "overseas":
        lines, errors = _parse_overseas(ws)
    else:
        raise ValueError(f"不认识的模板类型：{resolved_type}")

    return ParsedPlan(template_type=resolved_type, lines=lines, errors=errors)


def _parse_quantity(value) -> tuple[int, str | None]:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0, f"数量格式不对，读到的是「{value!r}」，不是数字"
    if value <= 0:
        return 0, f"数量必须是正数，读到的是 {value}"
    if not float(value).is_integer():
        return 0, f"数量不是整数：{value}"
    return int(value), None


def _parse_walmart(ws: Worksheet, header_row: int) -> tuple[list[PlanLine], list[str]]:
    cols = column_index_map(ws, header_row)
    idx = require_columns(cols, WALMART_HEADERS, "沃尔玛发货计划表")
    shop_col, sku_col, qty_col = idx["店铺"], idx["RK-SKU"], idx["预计发货数量"]

    lines: list[PlanLine] = []
    errors: list[str] = []
    last_zd: str | None = None

    for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        shop_val = row[shop_col - 1].value
        if shop_val is not None and str(shop_val).strip():
            last_zd = str(shop_val).strip()

        sku_val = row[sku_col - 1].value
        if sku_val is None or not str(sku_val).strip():
            continue
        sku = str(sku_val).strip()

        qty_val = row[qty_col - 1].value
        if qty_val is None or qty_val == "":
            continue

        if not last_zd:
            errors.append(f"第{row_no}行（SKU={sku}）：这一行之前都没有出现过店铺编号")
            continue

        qty, err = _parse_quantity(qty_val)
        if err:
            errors.append(f"第{row_no}行（SKU={sku}）：{err}")
            continue

        lines.append(
            PlanLine(
                zd=last_zd,
                sku_kind="RK",
                sku=sku,
                quantity=qty,
                destination_label=last_zd,
                source_row=row_no,
            )
        )

    return lines, errors


def _parse_amazon(ws: Worksheet, header_row: int) -> tuple[list[PlanLine], list[str]]:
    cols = column_index_map(ws, header_row)
    idx = require_columns(cols, AMAZON_HEADERS, "亚马逊发货计划表")
    shop_col, sku_col = idx["店铺"], idx["SKU"]
    dest_cols = [
        (str(name), col) for name, col in cols.items() if col not in (shop_col, sku_col)
    ]
    if not dest_cols:
        raise HeaderNotFoundError("亚马逊发货计划表：SKU 列右边没找到任何目的地数量列")

    lines: list[PlanLine] = []
    errors: list[str] = []
    last_zd: str | None = None

    for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        shop_val = row[shop_col - 1].value
        if shop_val is not None and str(shop_val).strip():
            last_zd = str(shop_val).strip()

        sku_val = row[sku_col - 1].value
        if sku_val is None or not str(sku_val).strip():
            continue
        sku = str(sku_val).strip()

        row_has_qty = False
        for dest_name, dest_col in dest_cols:
            qty_val = row[dest_col - 1].value
            if qty_val is None or qty_val == "":
                continue
            row_has_qty = True

            if not last_zd:
                errors.append(f"第{row_no}行（SKU={sku}，{dest_name}）：这一行之前都没有出现过店铺编号")
                continue

            qty, err = _parse_quantity(qty_val)
            if err:
                errors.append(f"第{row_no}行（SKU={sku}，{dest_name}）：{err}")
                continue

            lines.append(
                PlanLine(
                    zd=last_zd,
                    sku_kind="AMZ",
                    sku=sku,
                    quantity=qty,
                    destination_label=dest_name,
                    source_row=row_no,
                )
            )

        _ = row_has_qty  # 没有任何目的地填数量的行，直接跳过，不算错误（可能只是占位）

    return lines, errors


def _parse_overseas(ws: Worksheet) -> tuple[list[PlanLine], list[str]]:
    header_row = find_header_row(ws, OVERSEAS_HEADERS, max_scan_rows=3)
    cols = column_index_map(ws, header_row)
    idx = require_columns(cols, OVERSEAS_HEADERS, "海外仓发货计划表")
    sku_col = idx["海外仓-SKU"]

    dest_row = header_row + 1
    dest_cols = [
        (str(cell.value).strip(), cell.column)
        for cell in ws[dest_row]
        if cell.value is not None and cell.column > sku_col
    ]
    if not dest_cols:
        raise HeaderNotFoundError("海外仓发货计划表：表头下一行没找到任何目的仓列（ZD 编号）")

    lines: list[PlanLine] = []
    errors: list[str] = []

    for row_no, row in enumerate(ws.iter_rows(min_row=dest_row + 1), start=dest_row + 1):
        sku_val = row[sku_col - 1].value
        if sku_val is None or not str(sku_val).strip():
            continue
        sku = str(sku_val).strip()

        for zd, dest_col in dest_cols:
            qty_val = row[dest_col - 1].value
            if qty_val is None or qty_val == "":
                continue

            qty, err = _parse_quantity(qty_val)
            if err:
                errors.append(f"第{row_no}行（SKU={sku}，{zd}）：{err}")
                continue

            lines.append(
                PlanLine(
                    zd=zd,
                    sku_kind="RK",
                    sku=sku,
                    quantity=qty,
                    destination_label=zd,
                    source_row=row_no,
                )
            )

    return lines, errors
