"""采购订单汇总表：每一行是"一笔采购订单里的一个型号"，后面跟着一长串"日期列"（哪一批次
出货，就在那一列写发出的数量），"未出货数量"是表格里已经有的本地公式：
    =+<订单数量列><行>-SUM(<第一个日期列><行>:<最后一个日期列><行>)
本工具只需要把这次发货的数量填进正确的日期列——公式本身在 Excel 打开时会自己算出新的
"未出货数量"，不用我们手动改那一格。

但如果要写入的日期还没有对应的列，需要真的在表格中间插入一新列（按日期顺序插在合适的位置，
而不是随手加在最后，方便人工阅读）。openpyxl 的 insert_cols 只会把已有单元格的内容搬到新位置，
并不会像 Excel 那样自动把公式里的区间引用（比如 SUM(K4:BU4)）跟着调整——如果不管这件事，
插入一列之后所有行的"未出货数量"公式实际引用的区间就会跟数据错位，算出错的数字还看不出来。
所以每次插入日期列之后，都要把所有行的"未出货数量"公式重新按当前的实际列范围写一遍。

分摊规则：同一个货号（型号）可能对应好几笔采购订单，按采购日期从早到晚，先把早的写满，
写满了（未出货数量到 0）就换下一笔。如果找不到货号对应的订单，或者所有能找到的订单加起来
的未出货数量还是不够这次要发的数量，就是真的缺货，报 shortfall 让调用方报错，不找别的货号
顶替——"变体"字段标的是"同一造型不同颜色"的分组，不是库存可以互换的依据（核对过真实数据，
不同颜色的货没法互相顶替发货），所以分摊只在同一个货号自己名下的订单里找，这里只负责
"给定一个型号能分摊出多少、还差多少"。
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

from copy import copy

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .column_utils import HeaderNotFoundError, column_index_map, find_header_row, require_columns

FIXED_HEADERS = ["订单号", "采购日期", "型号", "订单数量", "数量单位", "未出货数量"]
SUB_HEADER_LABEL = "出货时间"

EXCEL_EPOCH = dt.datetime(1899, 12, 30)


def _to_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (EXCEL_EPOCH + dt.timedelta(days=value)).date()
        except (OverflowError, OSError):
            return None
    if isinstance(value, str):
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


@dataclass
class PurchaseRow:
    row_index: int
    order_no: str
    purchase_date: dt.date
    model: str
    order_qty: int
    initial_remaining: int
    consumed_this_run: int = 0

    @property
    def remaining(self) -> int:
        return self.initial_remaining - self.consumed_this_run


@dataclass
class Allocation:
    row: PurchaseRow
    quantity: int


@dataclass
class AllocationOutcome:
    allocations: list[Allocation] = field(default_factory=list)
    shortfall: int = 0


class PurchaseBook:
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.header_row = find_header_row(ws, FIXED_HEADERS, max_scan_rows=5)
        self.sub_header_row = self.header_row + 1

        cols = column_index_map(ws, self.header_row)
        idx = require_columns(cols, FIXED_HEADERS, "采购订单汇总表")
        self.order_no_col = idx["订单号"]
        self.purchase_date_col = idx["采购日期"]
        self.model_col = idx["型号"]
        self.order_qty_col = idx["订单数量"]
        self.remaining_col = idx["未出货数量"]

        self.date_col_start = idx["数量单位"] + 1
        self.date_col_end = self.remaining_col - 1
        if self.date_col_end < self.date_col_start:
            raise HeaderNotFoundError("采购订单汇总表：在「数量单位」和「未出货数量」之间没找到日期列")

        self.rows: list[PurchaseRow] = []
        self.by_model: dict[str, list[PurchaseRow]] = {}
        self._load_rows()

    # ---- 读取 ----

    def _load_rows(self) -> None:
        self.rows.clear()
        self.by_model.clear()
        last_row = self.ws.max_row
        for r in range(self.header_row + 2, last_row + 1):
            model = self.ws.cell(row=r, column=self.model_col).value
            if model is None or not str(model).strip():
                continue
            model = str(model).strip()

            order_no = self.ws.cell(row=r, column=self.order_no_col).value
            purchase_date = _to_date(self.ws.cell(row=r, column=self.purchase_date_col).value)
            order_qty = self.ws.cell(row=r, column=self.order_qty_col).value
            if purchase_date is None or not isinstance(order_qty, (int, float)):
                continue

            shipped_sum = 0
            for c in range(self.date_col_start, self.date_col_end + 1):
                v = self.ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    shipped_sum += v

            row_obj = PurchaseRow(
                row_index=r,
                order_no=str(order_no).strip() if order_no is not None else "",
                purchase_date=purchase_date,
                model=model,
                order_qty=int(order_qty),
                initial_remaining=int(order_qty) - int(shipped_sum),
            )
            self.rows.append(row_obj)
            self.by_model.setdefault(model, []).append(row_obj)

        for candidates in self.by_model.values():
            candidates.sort(key=lambda r: (r.purchase_date, r.row_index))

    # ---- 分摊 ----

    def allocate(self, model: str, quantity_needed: int) -> AllocationOutcome:
        outcome = AllocationOutcome()
        remaining_need = quantity_needed

        for row_obj in self.by_model.get(model, []):
            if remaining_need <= 0:
                break
            avail = row_obj.remaining
            if avail <= 0:
                continue
            take = min(avail, remaining_need)
            row_obj.consumed_this_run += take
            outcome.allocations.append(Allocation(row=row_obj, quantity=take))
            remaining_need -= take

        outcome.shortfall = remaining_need
        return outcome

    # ---- 写入 ----

    def _real_date_columns(self) -> list[tuple[int, dt.date | None]]:
        # SUM 区间（date_col_start..date_col_end）里混了几个不是真正"出货批次日期"的列
        # （比如表头是裸数字、行3副标题不是"出货时间"的几列），这些列依然要算在未出货数量的
        # 公式范围内（跟表格原有公式保持一致），但找"哪一列对应这个日期"、"新日期该插在哪"
        # 的时候，只能看行3副标题确实是"出货时间"的那些列，不然会被那几个裸数字表头误判成
        # 很晚/很早的日期。
        cols = []
        for c in range(self.date_col_start, self.date_col_end + 1):
            if self.ws.cell(row=self.sub_header_row, column=c).value == SUB_HEADER_LABEL:
                d = _to_date(self.ws.cell(row=self.header_row, column=c).value)
                cols.append((c, d))
        return cols

    def find_or_create_date_column(self, target_date: dt.date) -> int:
        real_cols = self._real_date_columns()

        for c, d in real_cols:
            if d == target_date:
                return c

        insert_at = (real_cols[-1][0] + 1) if real_cols else (self.date_col_end + 1)
        for c, d in real_cols:
            if d is not None and d > target_date:
                insert_at = c
                break

        # 插入之前先记下要照抄格式的那一列（本来在插入点左边的那一列，插入之后位置不变，
        # 还是原来的列号；插入点正好是日期区间最左边、左边没有可抄的日期列时，就抄插入之后
        # 落在右边的那一列）——旧列右移之后这个"左边列号"不会跟着变，插入完了正好能直接用。
        style_source = insert_at - 1 if insert_at - 1 >= self.date_col_start else None

        self.ws.insert_cols(insert_at)
        # insert_cols 只搬单元格本身，"整列"级别的设置（列宽等）和新插入的这一整列的格子样式
        # （字体/填充/边框），都不会像 Excel 那样自动跟着处理——列宽挂在列号上，列号错位了但
        # 设置没跟着挪；新插入的这一整列每一行都是没有任何格式的空白格子，直接看会很突兀。
        self._shift_column_dimensions(insert_at)
        self._copy_column_style(insert_at, style_source if style_source is not None else insert_at + 1)

        self.ws.cell(row=self.header_row, column=insert_at, value=dt.datetime.combine(target_date, dt.time()))
        self.ws.cell(row=self.sub_header_row, column=insert_at, value=SUB_HEADER_LABEL)

        self.date_col_end += 1
        if insert_at <= self.remaining_col:
            self.remaining_col += 1
        self._rewrite_remaining_formulas()

        return insert_at

    def _shift_column_dimensions(self, inserted_at: int) -> None:
        # 从最右边的列开始往左处理，不然后面的赋值会覆盖掉还没读出来的旧值（跟
        # shipment_summary.py 里 _shift_row_heights 是同一个道理，只是行列换了个方向）。
        existing_cols = sorted(
            (
                idx
                for letter in self.ws.column_dimensions
                if (idx := column_index_from_string(letter)) >= inserted_at
            ),
            reverse=True,
        )
        for idx in existing_cols:
            src = self.ws.column_dimensions[get_column_letter(idx)]
            dest = self.ws.column_dimensions[get_column_letter(idx + 1)]
            dest.width = src.width
            dest.hidden = src.hidden
            dest.outlineLevel = src.outlineLevel
            dest.collapsed = src.collapsed
        inserted_letter = get_column_letter(inserted_at)
        if inserted_letter in self.ws.column_dimensions:
            del self.ws.column_dimensions[inserted_letter]

    def _copy_column_style(self, dest_col: int, src_col: int) -> None:
        if self.ws.max_row is None:
            return
        src_letter = get_column_letter(src_col)
        if src_letter in self.ws.column_dimensions:
            src_dim = self.ws.column_dimensions[src_letter]
            dest_dim = self.ws.column_dimensions[get_column_letter(dest_col)]
            dest_dim.width = src_dim.width
        for r in range(1, self.ws.max_row + 1):
            src_cell = self.ws.cell(row=r, column=src_col)
            if not src_cell.has_style:
                continue
            dest_cell = self.ws.cell(row=r, column=dest_col)
            dest_cell.font = copy(src_cell.font)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.border = copy(src_cell.border)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.number_format = src_cell.number_format
            dest_cell.protection = copy(src_cell.protection)

    def _rewrite_remaining_formulas(self) -> None:
        qty_letter = get_column_letter(self.order_qty_col)
        start_letter = get_column_letter(self.date_col_start)
        end_letter = get_column_letter(self.date_col_end)
        for row_obj in self.rows:
            r = row_obj.row_index
            formula = f"=+{qty_letter}{r}-SUM({start_letter}{r}:{end_letter}{r})"
            self.ws.cell(row=r, column=self.remaining_col, value=formula)

    def write_allocation(self, allocation: Allocation, date_col: int) -> None:
        cell = self.ws.cell(row=allocation.row.row_index, column=date_col)
        existing = cell.value
        new_value = (existing or 0) + allocation.quantity if isinstance(existing, (int, float)) else allocation.quantity
        cell.value = new_value
