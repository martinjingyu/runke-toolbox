"""发货计划自动更新——这个工具的界面。

跟其它工具比，这个要小心得多：真的会改真实的采购汇总表和发货计划汇总表这两份数据。所以
界面分两步，中间必须停下来让人看："生成预览"只在内存里跑完整个流程，不落盘；确认没问题了
再点"确认写入"，这时候才备份原文件、真的存盘。

预览、写入都可能要处理几千上万行的表格（发货计划汇总表插入一行要重新扫描它后面所有行的
公式，见 shipment_summary.py），跑起来可能要几分钟，所以放在后台线程里跑，不能卡住界面。
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
from PySide6.QtCore import QDate, QSettings, Qt, QThread, Signal
from PySide6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from core.backup import backup_file
from core.diff_preview import DiffPreviewGroup

from .column_utils import column_index_map
from .diff import PreviewResult, run_and_capture_diff
from .planner import Plan, build_plan
from .product_lookup import load_product_lookup
from .purchase_book import PurchaseBook
from .shipment_summary import ShipmentSummaryBook
from .shipment_templates import PlanLine, list_sheet_names, parse_shipment_plan

_EDITABLE_FIELD = "备注"

_TEMPLATE_LABELS = {"walmart": "沃尔玛", "amazon": "亚马逊", "overseas": "海外仓"}
_LABEL_TO_TEMPLATE = {v: k for k, v in _TEMPLATE_LABELS.items()}
_AUTO_LABEL = "自动识别"

# 这三张表长期维护、路径基本不会变，记住上次选过的路径，下次打开软件自动填上，不用每次
# 都重新浏览——存到 QSettings 里（Windows 是注册表，Mac 是本地 ini 文件），跟软件本身的
# 数据文件（config.local.yaml、data/ 目录）无关，纯粹是这个输入框的"上次填了什么"。
_SETTINGS_KEY_PRODUCT = "shipment_plan_apply/product_path"
_SETTINGS_KEY_PURCHASE = "shipment_plan_apply/purchase_path"
_SETTINGS_KEY_SUMMARY = "shipment_plan_apply/summary_path"


def _next_tuesday(today: dt.date | None = None) -> dt.date:
    today = today or dt.date.today()
    days_ahead = (1 - today.weekday()) % 7  # Monday=0 ... Tuesday=1
    return today + dt.timedelta(days=days_ahead)


def _file_picker_row(label_text: str, on_browse) -> tuple[QHBoxLayout, QLineEdit]:
    row = QHBoxLayout()
    row.addWidget(QLabel(label_text))
    line_edit = QLineEdit()
    line_edit.setReadOnly(True)
    row.addWidget(line_edit, 1)
    button = QPushButton("浏览…")
    button.clicked.connect(on_browse)
    row.addWidget(button)
    return row, line_edit



class _PlanFileEntry:
    def __init__(self, path: str, sheet_names: list[str], detected: str | None):
        self.path = path
        self.sheet_names = sheet_names
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItems(sheet_names)
        self.template_combo = QComboBox()
        self.template_combo.addItem(_AUTO_LABEL)
        self.template_combo.addItems(list(_TEMPLATE_LABELS.values()))
        if detected in _TEMPLATE_LABELS:
            self.template_combo.setCurrentText(_TEMPLATE_LABELS[detected])

    @property
    def sheet_name(self) -> str:
        return self.sheet_combo.currentText()

    @property
    def template_type(self) -> str | None:
        label = self.template_combo.currentText()
        return _LABEL_TO_TEMPLATE.get(label)


class _PreviewWorker(QThread):
    succeeded = Signal(object)  # (plan, result_or_None, purchase_wb, summary_wb, purchase_book, summary_book)
    failed = Signal(str)
    progress = Signal(int, int)  # done, total（分摊记录数——真正耗时的插入行操作按这个报进度）

    def __init__(
        self,
        product_path: str,
        purchase_path: str,
        summary_path: str,
        plan_files: list[tuple[str, str, str | None]],  # (path, sheet_name, template_type)
        ship_date: dt.date,
    ):
        super().__init__()
        self._product_path = product_path
        self._purchase_path = purchase_path
        self._summary_path = summary_path
        self._plan_files = plan_files
        self._ship_date = ship_date

    def run(self):
        try:
            lookup = load_product_lookup(Path(self._product_path))

            all_lines: list[PlanLine] = []
            parse_errors: list[str] = []
            for path, sheet_name, template_type in self._plan_files:
                parsed = parse_shipment_plan(Path(path), sheet_name, template_type)
                file_label = Path(path).name
                for line in parsed.lines:
                    line.source_file = file_label
                all_lines.extend(parsed.lines)
                parse_errors.extend(f"[{file_label}] {e}" for e in parsed.errors)

            purchase_wb = openpyxl.load_workbook(self._purchase_path, data_only=False)
            purchase_book = PurchaseBook(purchase_wb.active)
            summary_wb = openpyxl.load_workbook(self._summary_path, data_only=False)
            summary_book = ShipmentSummaryBook(summary_wb.active)

            plan = build_plan(all_lines, parse_errors, lookup, purchase_book, self._ship_date)

            if plan.has_blocking_errors:
                self.succeeded.emit((plan, None, purchase_wb, summary_wb, purchase_book, summary_book))
                return

            result = run_and_capture_diff(
                plan,
                purchase_book,
                summary_book,
                progress_callback=lambda done, total: self.progress.emit(done, total),
            )
            self.succeeded.emit((plan, result, purchase_wb, summary_wb, purchase_book, summary_book))
        except Exception as exc:
            self.failed.emit(str(exc))


class ShipmentPlanApplyPanel(QWidget):
    def __init__(self):
        super().__init__()
        self._plan_entries: list[_PlanFileEntry] = []
        self._worker: _PreviewWorker | None = None
        self._settings = QSettings()

        self._plan: Plan | None = None
        self._result: PreviewResult | None = None
        self._purchase_wb = None
        self._summary_wb = None
        self._purchase_path = ""
        self._summary_path = ""
        self._purchase_header_row = None
        self._summary_header_row = None

        # 内容可能很长（预览表格行数不定），整个面板放进一个纵向可滚动的区域里，而不是让每个
        # 子控件（比如表格）各自滚动——外层容器统一上下滚，体验更接近正常网页/文档。
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        outer_layout.addWidget(scroll_area)

        content = QWidget()
        scroll_area.setWidget(content)
        layout = QVBoxLayout(content)

        title = QLabel("发货计划自动更新")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(title)
        note = QLabel(
            "会真的修改采购订单汇总表和发货计划汇总表——请先点「生成预览」检查没问题，再点「确认写入」。"
            "确认写入前会自动备份这两份文件的原始版本。"
        )
        note.setWordWrap(True)
        layout.addWidget(note)

        inputs_box = QGroupBox("长期维护的三张表")
        inputs_layout = QVBoxLayout(inputs_box)
        row, self._product_edit = _file_picker_row("在售产品信息总表", self._browse_product)
        inputs_layout.addLayout(row)
        row, self._purchase_edit = _file_picker_row("采购订单汇总表", self._browse_purchase)
        inputs_layout.addLayout(row)
        row, self._summary_edit = _file_picker_row("发货计划汇总表", self._browse_summary)
        inputs_layout.addLayout(row)
        layout.addWidget(inputs_box)

        self._product_edit.setText(self._settings.value(_SETTINGS_KEY_PRODUCT, ""))
        self._purchase_edit.setText(self._settings.value(_SETTINGS_KEY_PURCHASE, ""))
        self._summary_edit.setText(self._settings.value(_SETTINGS_KEY_SUMMARY, ""))

        plan_box = QGroupBox("运营提供的发货计划表（可以一次导入好几份）")
        plan_layout = QVBoxLayout(plan_box)
        add_row = QHBoxLayout()
        add_button = QPushButton("添加文件…")
        add_button.clicked.connect(self._add_plan_files)
        add_row.addWidget(add_button)
        add_row.addStretch(1)
        plan_layout.addLayout(add_row)

        self._plan_table = QTableWidget(0, 4)
        self._plan_table.setHorizontalHeaderLabels(["文件", "选哪个 sheet", "模板类型", ""])
        self._plan_table.horizontalHeader().setStretchLastSection(False)
        self._plan_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        plan_layout.addWidget(self._plan_table)
        layout.addWidget(plan_box)

        date_row = QHBoxLayout()
        date_row.addWidget(QLabel("这次发货日期（采购汇总表写进这一列；发货计划汇总表的发货时间也用这个）"))
        self._date_edit = QDateEdit()
        self._date_edit.setCalendarPopup(True)
        default_date = _next_tuesday()
        self._date_edit.setDate(QDate(default_date.year, default_date.month, default_date.day))
        date_row.addWidget(self._date_edit)
        date_row.addStretch(1)
        layout.addLayout(date_row)

        run_row = QHBoxLayout()
        self._preview_button = QPushButton("生成预览")
        self._preview_button.clicked.connect(self._start_preview)
        run_row.addWidget(self._preview_button)
        self._confirm_button = QPushButton("确认写入")
        self._confirm_button.setEnabled(False)
        self._confirm_button.clicked.connect(self._confirm_write)
        run_row.addWidget(self._confirm_button)
        run_row.addStretch(1)
        layout.addLayout(run_row)

        self._progress = QProgressBar()
        self._progress.setRange(0, 0)
        self._progress.hide()
        layout.addWidget(self._progress)

        self._status_label = QLabel("")
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)

        self._error_text = QTextEdit()
        self._error_text.setReadOnly(True)
        self._error_text.hide()
        self._error_text.setMaximumHeight(160)
        layout.addWidget(self._error_text)

        # 预览就两张表，一张表一整行：红色是变化前、绿色是变化后，同一笔记录的前后状态紧挨着，
        # 不用再左右对照着看。每张表上面有个筛选框，表里只有"备注"这一格能编辑，改了直接写回
        # 对应的工作表（不用等"确认写入"再收集一遍）。这个组件是通用的，见 core/diff_preview.py。
        self._purchase_diff_group = DiffPreviewGroup(
            "采购订单汇总表 · 改动对比", key_fields=["订单号", "型号"], editable_field=_EDITABLE_FIELD
        )
        layout.addWidget(self._purchase_diff_group)
        self._summary_diff_group = DiffPreviewGroup(
            "发货计划汇总表 · 改动对比", key_fields=["采购单号", "型号"], editable_field=_EDITABLE_FIELD
        )
        layout.addWidget(self._summary_diff_group)

    # ---- 文件选择 ----

    def _browse_product(self):
        start_dir = str(Path(self._product_edit.text()).parent) if self._product_edit.text() else ""
        path, _ = QFileDialog.getOpenFileName(self, "选择在售产品信息总表", start_dir, "Excel 文件 (*.xlsx)")
        if path:
            self._product_edit.setText(path)
            self._settings.setValue(_SETTINGS_KEY_PRODUCT, path)

    def _browse_purchase(self):
        start_dir = str(Path(self._purchase_edit.text()).parent) if self._purchase_edit.text() else ""
        path, _ = QFileDialog.getOpenFileName(self, "选择采购订单汇总表", start_dir, "Excel 文件 (*.xlsx)")
        if path:
            self._purchase_edit.setText(path)
            self._settings.setValue(_SETTINGS_KEY_PURCHASE, path)

    def _browse_summary(self):
        start_dir = str(Path(self._summary_edit.text()).parent) if self._summary_edit.text() else ""
        path, _ = QFileDialog.getOpenFileName(self, "选择发货计划汇总表", start_dir, "Excel 文件 (*.xlsx)")
        if path:
            self._summary_edit.setText(path)
            self._settings.setValue(_SETTINGS_KEY_SUMMARY, path)

    def _add_plan_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择运营提供的发货计划表", "", "Excel 文件 (*.xlsx)")
        for path in paths:
            try:
                sheet_names = list_sheet_names(Path(path))
                detected = None
                if sheet_names:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    from .shipment_templates import detect_template_type

                    detected = detect_template_type(wb[sheet_names[0]])
            except Exception as exc:
                QMessageBox.warning(self, "读取失败", f"「{Path(path).name}」读取失败：{exc}")
                continue

            entry = _PlanFileEntry(path, sheet_names, detected)
            self._plan_entries.append(entry)
            self._append_plan_row(entry)

    def _append_plan_row(self, entry: _PlanFileEntry) -> None:
        row = self._plan_table.rowCount()
        self._plan_table.insertRow(row)
        self._plan_table.setItem(row, 0, QTableWidgetItem(Path(entry.path).name))
        self._plan_table.setCellWidget(row, 1, entry.sheet_combo)
        self._plan_table.setCellWidget(row, 2, entry.template_combo)
        remove_button = QPushButton("移除")
        remove_button.clicked.connect(lambda: self._remove_plan_entry(entry))
        self._plan_table.setCellWidget(row, 3, remove_button)
        self._plan_table.resizeColumnsToContents()

    def _remove_plan_entry(self, entry: _PlanFileEntry) -> None:
        if entry not in self._plan_entries:
            return
        idx = self._plan_entries.index(entry)
        self._plan_entries.pop(idx)
        self._plan_table.removeRow(idx)

    # ---- 预览 ----

    def _start_preview(self):
        product_path = self._product_edit.text().strip()
        purchase_path = self._purchase_edit.text().strip()
        summary_path = self._summary_edit.text().strip()

        missing = []
        if not product_path:
            missing.append("在售产品信息总表")
        if not purchase_path:
            missing.append("采购订单汇总表")
        if not summary_path:
            missing.append("发货计划汇总表")
        if not self._plan_entries:
            missing.append("发货计划表（至少一份）")
        if missing:
            QMessageBox.warning(self, "缺少输入", "还没选：" + "、".join(missing))
            return

        qdate = self._date_edit.date()
        ship_date = dt.date(qdate.year(), qdate.month(), qdate.day())

        plan_files = [(e.path, e.sheet_name, e.template_type) for e in self._plan_entries]

        self._preview_button.setEnabled(False)
        self._confirm_button.setEnabled(False)
        self._progress.setRange(0, 0)  # 分摊笔数还没算出来之前先显示忙碌样式，算出来后会切成百分比
        self._progress.show()
        self._error_text.hide()
        self._status_label.setText("正在处理……表格行数多的话可能要几分钟，请耐心等待，不要关闭窗口。")
        self._clear_preview_tables()

        self._purchase_path = purchase_path
        self._summary_path = summary_path

        self._worker = _PreviewWorker(product_path, purchase_path, summary_path, plan_files, ship_date)
        self._worker.succeeded.connect(self._on_preview_succeeded)
        self._worker.failed.connect(self._on_preview_failed)
        self._worker.progress.connect(self._on_progress)
        self._worker.start()

    def _on_progress(self, done: int, total: int):
        if total <= 0:
            return
        if self._progress.maximum() != total:
            self._progress.setRange(0, total)
        self._progress.setValue(done)
        self._status_label.setText(f"正在写入变化：{done}/{total} 笔分摊……")

    def _clear_preview_tables(self):
        self._purchase_diff_group.clear()
        self._summary_diff_group.clear()

    def _on_preview_succeeded(self, payload):
        plan, result, purchase_wb, summary_wb, purchase_book, summary_book = payload
        self._progress.hide()
        self._preview_button.setEnabled(True)
        self._plan = plan
        self._result = result
        self._purchase_wb = purchase_wb
        self._summary_wb = summary_wb
        self._purchase_header_row = purchase_book.header_row
        self._summary_header_row = summary_book.header_row

        if plan.has_blocking_errors:
            self._status_label.setText(
                f"这一批有 {len(plan.parse_errors) + sum(1 for i in plan.items if i.errors)} 处问题，"
                "全部列在下面，一个都没写入。"
            )
            lines = list(plan.parse_errors)
            for item in plan.items:
                lines.extend(item.errors)
            self._error_text.setPlainText("\n".join(lines))
            self._error_text.show()
            self._confirm_button.setEnabled(False)
            return

        self._error_text.hide()
        self._status_label.setText(
            f"预览完成，共 {len(plan.items)} 条记录，涉及 {plan.total_allocations} 笔采购订单分摊。"
            "确认没问题的话点「确认写入」。"
        )
        self._fill_preview(result)
        self._confirm_button.setEnabled(True)
        self._confirm_button.setText("确认写入")

    def _fill_preview(self, result: PreviewResult):
        purchase_ws = self._purchase_wb.active
        summary_ws = self._summary_wb.active
        purchase_remark_col = column_index_map(purchase_ws, self._purchase_header_row).get(_EDITABLE_FIELD)
        summary_remark_col = column_index_map(summary_ws, self._summary_header_row).get(_EDITABLE_FIELD)

        self._purchase_diff_group.fill(result.purchase, ws=purchase_ws, col_index=purchase_remark_col)
        self._summary_diff_group.fill(result.summary, ws=summary_ws, col_index=summary_remark_col)

    def _on_preview_failed(self, message: str):
        self._progress.hide()
        self._preview_button.setEnabled(True)
        self._status_label.setText("预览失败，见弹窗说明。")
        QMessageBox.critical(self, "预览失败", message)

    # ---- 确认写入 ----

    def _confirm_write(self):
        if self._plan is None or self._plan.has_blocking_errors or self._purchase_wb is None:
            return

        reply = QMessageBox.question(
            self,
            "确认写入",
            "确定要把上面预览的改动写进：\n"
            f"  {self._purchase_path}\n"
            f"  {self._summary_path}\n\n"
            "写入前会先在原文件旁边自动生成一份备份。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            purchase_backup = backup_file(self._purchase_path)
            summary_backup = backup_file(self._summary_path)
        except Exception as exc:
            QMessageBox.critical(self, "备份失败", f"没能先备份原文件，写入已取消：{exc}")
            return

        try:
            self._purchase_wb.save(self._purchase_path)
            self._summary_wb.save(self._summary_path)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "写入失败",
                f"存盘失败：{exc}\n\n备份文件还在（{purchase_backup} / {summary_backup}），原文件可能已经部分改动，建议手动检查。",
            )
            return

        self._status_label.setText(
            f"已写入。备份文件：{purchase_backup.name} / {summary_backup.name}"
        )
        self._confirm_button.setEnabled(False)
        self._confirm_button.setText("已写入")

    def stop_running_tasks(self):
        if self._worker is not None and self._worker.isRunning():
            self._worker.wait(5000)
