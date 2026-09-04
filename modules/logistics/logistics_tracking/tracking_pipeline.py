"""读物流跟踪表格，按"货件状态"筛出还没完成的运单，按物流商分组、并行去各个货代平台查最后
路由，算出预览（不落盘）；人工确认没问题之后，调用方（panel.py）才把结果真的写回
"最后流水"/"是否有更新"两列、存盘。改自独立脚本 excel_pipeline.py，原来那份脚本只认写死的
一个文件路径 + 一个 sheet 名字，这里把这两样都变成调用方传入的参数，具体查询逻辑挪到
platforms/ 下（见 platforms/registry.py）。

跑的范围：只看"货件状态"不含"完成"字样（或者本身是空）的行——已经签收完成的运单路由不会
再变了，没必要每次都重新查一遍，这条规则原样照搬自独立脚本（当时跟业务确认过）。

"是否有更新"这一列的判断：新查到的"最后流水"如果跟这一行原来就有的"最后流水"不一样（包括
原来是空的情况），才算"有更新"；查询本身失败的话，直接把失败原因写进去（未接入自动查询/
还没配账号/未找到该运单/暂无路由信息等），不是留空，方便一眼看出是哪个环节卡住的。
"""
from __future__ import annotations

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .platforms.registry import Account, get_last_routes_for_carrier

REQUIRED_HEADERS = ["物流商", "货件状态"]
# 不同年份的表（"2026年SK出货跟踪表" vs "2025年RK出货跟踪总表"）这一列的表头文字不完全一样，
# 拿真实数据核对过：语义是同一个字段（运单号），按顺序试这几个都没有就真的报错，不去猜别的列。
WAYBILL_HEADER_CANDIDATES = ["运单号（必填）", "运单号", "物流单号"]
COL_LAST_ROUTE_HEADER = "最后流水"
COL_HAS_UPDATE_HEADER = "是否有更新"


class HeaderNotFoundError(Exception):
    pass


def list_sheet_names(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def _find_header_row(ws: Worksheet, max_scan_rows: int = 5) -> int:
    # 按表格实际行数封顶，不然 ws[row_idx] 在 read_only 模式下超出 max_row 会直接抛 IndexError
    # 而不是返回空行——这个坑 shipment_plan_apply 那边已经踩过一次，见 CLAUDE.md。
    last_row = min(max_scan_rows, ws.max_row or 0)
    for row_idx in range(1, last_row + 1):
        values = {cell.value for cell in ws[row_idx]}
        if all(h in values for h in REQUIRED_HEADERS):
            return row_idx
    raise HeaderNotFoundError(f"在前 {max_scan_rows} 行里没找到包含表头 {REQUIRED_HEADERS} 的行")


def _column_index_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is not None and cell.value not in mapping:
            mapping[cell.value] = cell.column
    return mapping


def _in_scope(status) -> bool:
    return status is None or "完成" not in str(status)


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


@dataclass
class ScanRow:
    row: int
    carrier: str
    waybill: str
    status: object


@dataclass
class TrackingPreviewRow:
    row: int
    carrier: str
    waybill: str
    status: object
    old_last_route: object
    old_has_update: object
    new_last_route: str | None
    new_has_update: str


class TrackingSheet:
    """包一层，把"哪行是表头、每列在哪、还没有的列该插在哪"这些跟具体工作表结构相关的细节
    收在一起——跟 shipment_plan_apply 里 PurchaseBook/ShipmentSummaryBook 是同一个思路。
    """

    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.header_row = _find_header_row(ws)
        cols = _column_index_map(ws, self.header_row)

        missing = [h for h in REQUIRED_HEADERS if h not in cols]
        if missing:
            raise HeaderNotFoundError(f"物流跟踪表格缺少必须的表头：{missing}")
        self.carrier_col = cols["物流商"]
        self.status_col = cols["货件状态"]

        waybill_col = next((cols[h] for h in WAYBILL_HEADER_CANDIDATES if h in cols), None)
        if waybill_col is None:
            raise HeaderNotFoundError(
                f"物流跟踪表格缺少运单号列，试过这些表头都没找到：{WAYBILL_HEADER_CANDIDATES}"
            )
        self.waybill_col = waybill_col

        # 这两列可能还不存在（第一次对这张表跑）——先记下现状，真的要写的时候（apply_preview）
        # 再决定要不要追加新列，预览阶段只读不写。
        self.last_route_col = cols.get(COL_LAST_ROUTE_HEADER)
        self.has_update_col = cols.get(COL_HAS_UPDATE_HEADER)

    # ---- 读取 / 扫描 ----

    def scan_rows(self) -> list[ScanRow]:
        rows: list[ScanRow] = []
        last_row = self.ws.max_row or 0
        for r in range(self.header_row + 1, last_row + 1):
            carrier = self.ws.cell(row=r, column=self.carrier_col).value
            waybill = self.ws.cell(row=r, column=self.waybill_col).value
            status = self.ws.cell(row=r, column=self.status_col).value
            if not carrier or not waybill:
                continue
            if not _in_scope(status):
                continue
            rows.append(
                ScanRow(row=r, carrier=_text(carrier), waybill=_text(waybill), status=status)
            )
        return rows

    def _old_value(self, row: int, col: int | None):
        if col is None:
            return None
        return self.ws.cell(row=row, column=col).value

    def build_preview(
        self,
        accounts_by_platform: dict[str, list[Account]],
        progress_callback=None,
        max_workers: int | None = None,
    ) -> list[TrackingPreviewRow]:
        """先扫一遍表格，按物流商把要查的运单号分组，再并行地一个物流商一个物流商去查
        （concurrent.futures.ThreadPoolExecutor）——物流商之间互相独立，谁先查完不影响别人，
        没必要排队等。查询结果按运单号在本地精确匹配，不做模糊匹配（这份工具会直接改真实的
        跟踪表格，不能猜）。

        progress_callback(done, total)：total 是这一批涉及的物流商个数，done 是已经查完的
        物流商个数——粒度按"物流商"而不是"运单号"，因为一个物流商内部（比如批量导出）本身
        就是一次网络请求，没法再拆更细的进度。
        """
        rows = self.scan_rows()

        waybills_by_carrier: dict[str, set[str]] = defaultdict(set)
        for row in rows:
            waybills_by_carrier[row.carrier].add(row.waybill)

        route_by_carrier_waybill: dict[tuple[str, str], object] = {}
        total = len(waybills_by_carrier)
        done = 0
        if progress_callback is not None:
            progress_callback(done, total)

        if waybills_by_carrier:
            with ThreadPoolExecutor(max_workers=max_workers or total) as executor:
                future_to_carrier = {
                    executor.submit(
                        get_last_routes_for_carrier,
                        carrier,
                        accounts_by_platform.get(carrier, []),
                        sorted(waybills),
                    ): carrier
                    for carrier, waybills in waybills_by_carrier.items()
                }
                for future in as_completed(future_to_carrier):
                    carrier = future_to_carrier[future]
                    for wb, res in future.result().items():
                        route_by_carrier_waybill[(carrier, wb)] = res
                    done += 1
                    if progress_callback is not None:
                        progress_callback(done, total)

        previews: list[TrackingPreviewRow] = []
        for row in rows:
            old_last_route = self._old_value(row.row, self.last_route_col)
            old_has_update = self._old_value(row.row, self.has_update_col)
            res = route_by_carrier_waybill.get((row.carrier, row.waybill))

            if res is None:
                previews.append(
                    TrackingPreviewRow(
                        row=row.row, carrier=row.carrier, waybill=row.waybill, status=row.status,
                        old_last_route=old_last_route, old_has_update=old_has_update,
                        new_last_route=None, new_has_update="未查询(内部错误)",
                    )
                )
            elif res.last_route:
                has_update = "有更新" if _text(old_last_route) != _text(res.last_route) else "无更新"
                previews.append(
                    TrackingPreviewRow(
                        row=row.row, carrier=row.carrier, waybill=row.waybill, status=row.status,
                        old_last_route=old_last_route, old_has_update=old_has_update,
                        new_last_route=res.last_route, new_has_update=has_update,
                    )
                )
            else:
                previews.append(
                    TrackingPreviewRow(
                        row=row.row, carrier=row.carrier, waybill=row.waybill, status=row.status,
                        old_last_route=old_last_route, old_has_update=old_has_update,
                        new_last_route=None, new_has_update=res.error or "查询失败(未知原因)",
                    )
                )
        return previews

    # ---- 写入 ----

    def _find_or_append_column(self, header_text: str) -> int:
        """表头已经存在就用那一列；不存在就在最后一个有内容的表头列后面追加一列，并把这一整列
        的格式（字体/填充/边框/列宽）照抄紧挨着左边那一列——新插入的列 openpyxl 默认不带任何
        格式，直接看会很突兀，这个坑 shipment_plan_apply 那边已经踩过一次，见 CLAUDE.md。
        """
        cols = _column_index_map(self.ws, self.header_row)
        if header_text in cols:
            return cols[header_text]

        max_used_col = self.ws.max_column or 0
        new_col = max_used_col + 1
        self.ws.cell(row=self.header_row, column=new_col, value=header_text)
        if max_used_col >= 1:
            self._copy_column_style(new_col, max_used_col)
        return new_col

    def _copy_column_style(self, dest_col: int, src_col: int) -> None:
        if self.ws.max_row is None:
            return
        src_letter = get_column_letter(src_col)
        if src_letter in self.ws.column_dimensions:
            dest_dim = self.ws.column_dimensions[get_column_letter(dest_col)]
            dest_dim.width = self.ws.column_dimensions[src_letter].width
        for r in range(1, self.ws.max_row + 1):
            src_cell = self.ws.cell(row=r, column=src_col)
            if not src_cell.has_style:
                continue
            dest_cell = self.ws.cell(row=r, column=dest_col)
            dest_cell.font = copy(src_cell.font)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.border = copy(src_cell.border)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.number_format = src_cell.number_format
            dest_cell.protection = copy(src_cell.protection)

    def apply_preview(self, previews: list[TrackingPreviewRow]) -> None:
        """把预览结果写回工作表（最后流水/是否有更新两列，不存在就追加）。只改内存里的
        openpyxl Workbook 对象，落不落盘、要不要先备份，交给调用方（panel.py 的"确认写入"）。
        """
        if self.last_route_col is None:
            self.last_route_col = self._find_or_append_column(COL_LAST_ROUTE_HEADER)
        if self.has_update_col is None:
            self.has_update_col = self._find_or_append_column(COL_HAS_UPDATE_HEADER)

        for p in previews:
            if p.new_last_route:
                self.ws.cell(row=p.row, column=self.last_route_col, value=p.new_last_route)
            self.ws.cell(row=p.row, column=self.has_update_col, value=p.new_has_update)
