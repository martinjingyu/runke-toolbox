"""核对入口：三个输入——箱唛 PDF（一个仓一个文件）、翻译表（WM-SKU→货号）、发货计划表。

匹配链路：
    箱唛条码 -> GTIN + QUANTITY + SHIPMENT ID（可靠，条码给的）
    箱唛印刷文字 -> OCR 出 WM-SKU（同一 GTIN 的箱子只 OCR 一次代表页）
    WM-SKU -> 翻译表查出货号（原sku）
    (SHIPMENT ID, 货号) -> 发货计划表里查计划数量

用 SHIPMENT ID 而不是"仓库"做匹配 key 的一部分，是因为发货计划表里同一个 SKU+仓库会横跨
几十条不同批次的历史发货记录，按"SKU+仓库"加总会把好几个月的货全加在一起，跟当前这一批完全
对不上；SHIPMENT ID 就是箱唛上那批货的唯一标识，用它才能精确定位到"这一批"。
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import fitz

from .box_label import BoxLabel, ocr_single_sku, parse_box_labels, require_tesseract
from .shipping_plan import parse_shipping_plan
from .translation import normalize_sku, parse_translation


@dataclass
class ReconcileResult:
    sku: str  # 货号，原始写法（报表展示用）
    warehouse: str  # 这批箱子实际来自哪个仓库文件（可能不止一个）
    shipment_id: str
    planned_quantity: int
    actual_quantity: int
    box_count: int
    in_plan: bool
    match: bool


@dataclass
class RunReport:
    results: list[ReconcileResult]
    skipped_pages: dict[str, int]  # 输入文件名 -> 跳过的非箱标签页数（比如托盘汇总页）
    split_pdf_paths: list[Path]
    unresolved_gtins: set[str]  # OCR 读不出来、或者翻译表查不到对应货号的 GTIN


def run(
    pdf_paths: list[str | Path],
    translation_path: str | Path,
    plan_path: str | Path,
    output_dir: str | Path,
) -> RunReport:
    require_tesseract()

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wm_to_rk = parse_translation(translation_path)
    plan_rows = parse_shipping_plan(plan_path)

    # (SHIPMENT ID, 归一化后的货号) -> 计划数量
    planned_by_key: dict[tuple[str, str], int] = defaultdict(int)
    plan_display_sku: dict[str, str] = {}  # 归一化货号 -> 计划表里的原始写法，报表展示用
    for row in plan_rows:
        if not row.tracking_id:
            continue
        norm_sku = normalize_sku(row.sku)
        planned_by_key[(row.tracking_id, norm_sku)] += row.planned_quantity
        plan_display_sku.setdefault(norm_sku, row.sku)

    # gtin -> 仓库(=输入文件名，不含扩展名) -> [(pdf路径, BoxLabel), ...]
    by_gtin_warehouse: dict[str, dict[str, list[tuple[Path, BoxLabel]]]] = defaultdict(lambda: defaultdict(list))
    skipped_pages: dict[str, int] = {}
    for pdf_path in pdf_paths:
        pdf_path = Path(pdf_path)
        warehouse = pdf_path.stem
        labels, skipped = parse_box_labels(pdf_path)
        skipped_pages[pdf_path.name] = skipped
        for label in labels:
            by_gtin_warehouse[label.gtin][warehouse].append((pdf_path, label))

    # 每个 GTIN 只 OCR 一次代表页（同一个 GTIN 的箱子必然是同一个 SKU），翻译成货号；
    # OCR 失败或翻译表查不到，就记下来提醒用户，不静默丢弃
    gtin_to_sku: dict[str, str] = {}
    unresolved_gtins: set[str] = set()
    for gtin, by_warehouse in by_gtin_warehouse.items():
        first_pdf_path, first_label = next(iter(by_warehouse.values()))[0]
        wm_sku_raw = ocr_single_sku(first_pdf_path, first_label.page_index)
        rk_sku = wm_to_rk.get(normalize_sku(wm_sku_raw)) if wm_sku_raw else None
        if rk_sku is None:
            unresolved_gtins.add(gtin)
        gtin_to_sku[gtin] = rk_sku or wm_sku_raw or gtin

    split_pdf_paths = _write_split_pdfs(by_gtin_warehouse, gtin_to_sku, output_dir)

    # 按 (SHIPMENT ID, 归一化货号) 汇总实际数量
    actual_by_key: dict[tuple[str, str], int] = defaultdict(int)
    boxes_by_key: dict[tuple[str, str], int] = defaultdict(int)
    warehouses_by_key: dict[tuple[str, str], set[str]] = defaultdict(set)
    for gtin, by_warehouse in by_gtin_warehouse.items():
        norm_sku = normalize_sku(gtin_to_sku[gtin])
        for warehouse, entries in by_warehouse.items():
            for _, label in entries:
                key = (label.shipment_id, norm_sku)
                actual_by_key[key] += label.quantity
                boxes_by_key[key] += 1
                warehouses_by_key[key].add(warehouse)

    # 只看箱唛里实际出现过的 (SHIPMENT ID, 货号)——发货计划表里还有几万条跟这批货无关的历史记录，
    # 不是这次要核对的范围
    results = []
    for key in sorted(actual_by_key):
        shipment_id, norm_sku = key
        results.append(
            ReconcileResult(
                sku=plan_display_sku.get(norm_sku, norm_sku),
                warehouse="/".join(sorted(warehouses_by_key[key])),
                shipment_id=shipment_id,
                planned_quantity=planned_by_key.get(key, 0),
                actual_quantity=actual_by_key[key],
                box_count=boxes_by_key[key],
                in_plan=key in planned_by_key,
                match=(key in planned_by_key) and (planned_by_key[key] == actual_by_key[key]),
            )
        )

    return RunReport(
        results=results,
        skipped_pages=skipped_pages,
        split_pdf_paths=split_pdf_paths,
        unresolved_gtins=unresolved_gtins,
    )


def _write_split_pdfs(by_gtin_warehouse, gtin_to_sku: dict[str, str], output_dir: Path) -> list[Path]:
    split_pdf_paths: list[Path] = []
    for gtin, by_warehouse in by_gtin_warehouse.items():
        sku_name = gtin_to_sku[gtin]
        for warehouse, entries in by_warehouse.items():
            out_doc = fitz.open()
            src_cache: dict[Path, fitz.Document] = {}
            try:
                for pdf_path, label in entries:
                    if pdf_path not in src_cache:
                        src_cache[pdf_path] = fitz.open(pdf_path)
                    out_doc.insert_pdf(src_cache[pdf_path], from_page=label.page_index, to_page=label.page_index)
                out_path = output_dir / f"{_sanitize_filename(sku_name)}_{warehouse}.pdf"
                out_doc.save(out_path)
                split_pdf_paths.append(out_path)
            finally:
                out_doc.close()
                for src in src_cache.values():
                    src.close()
    return split_pdf_paths


def _sanitize_filename(name: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in "-_." else "_" for c in name)
    return cleaned or "unknown"


def write_report_xlsx(report: RunReport, output_path: str | Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "核对结果"

    headers = ["货号(SKU)", "仓库", "SHIPMENT ID", "计划数量", "实际数量", "箱数", "是否一致", "备注"]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ok_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    bad_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for r, item in enumerate(report.results, 2):
        note = "" if item.in_plan else "发货计划表里查不到这个 SHIPMENT ID + 货号的组合"
        row = [
            item.sku, item.warehouse, item.shipment_id,
            item.planned_quantity, item.actual_quantity, item.box_count,
            "一致" if item.match else "不一致", note,
        ]
        for col, value in enumerate(row, 1):
            ws.cell(row=r, column=col, value=value)
        fill = ok_fill if item.match else bad_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).fill = fill

    for col, width in zip("ABCDEFGH", [16, 14, 16, 10, 10, 8, 10, 34]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    wb.save(output_path)
