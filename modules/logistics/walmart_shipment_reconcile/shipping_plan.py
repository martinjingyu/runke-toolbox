"""读取发货计划表（比如 发货计划表.xlsx 5.18 .xlsx1.xlsx）。

表头不在第一行——前面几行是"出货单号/出货日期/物流渠道"这种表单式信息，真正的表头
（标签/仓库/数量……）在往下第几行不一定，所以自动扫描找到那一行，不写死行号。

同一个 SKU 在这张表里会跨很多个不同批次的发货记录（同一个 SKU + 同一个仓库能出现几十行，
横跨不同日期），不能直接按"SKU+仓库"加总来当计划数量——那样会把好几个月的历史批次全加在一起，
跟当前这一批货完全对不上。好在"追踪编号"这一列的值实测就是箱唛上的 SHIPMENT ID（核对过 69 条
记录完全对得上），所以真正能唯一定位"这一批货"的 key 是 (追踪编号, 标签)，不是 (仓库, 标签)。
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

_REQUIRED_HEADERS = ["标签", "仓库", "数量", "追踪编号"]
_WAREHOUSE_CODE_PATTERN = re.compile(r"[A-Z]{2,4}\d{1,2}")


@dataclass
class ShippingPlanRow:
    sku: str  # "标签"列，货号，保留原始写法
    warehouse_raw: str
    warehouse_code: str | None  # 归一化后的仓库代码，比如 "US(IND3)" -> "IND3"；识别不出来是 None
    tracking_id: str  # "追踪编号"，对应箱唛上的 SHIPMENT ID
    planned_quantity: int


def normalize_warehouse(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    m = _WAREHOUSE_CODE_PATTERN.search(text)
    return m.group(0) if m else None


def _find_header_row(rows: list[tuple]) -> tuple[int, list[str]]:
    for i, row in enumerate(rows):
        texts = [str(v).strip() if v is not None else "" for v in row]
        if all(req in texts for req in _REQUIRED_HEADERS):
            return i, texts
    raise ValueError(f"发货计划表里找不到包含 {_REQUIRED_HEADERS} 的表头行")


def parse_shipping_plan(xlsx_path: str | Path, sheet_name: str | None = None) -> list[ShippingPlanRow]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_idx, header = _find_header_row(rows)
    col = {name: header.index(name) for name in _REQUIRED_HEADERS}

    results = []
    for row in rows[header_idx + 1 :]:
        sku = row[col["标签"]]
        if sku is None:
            continue
        tracking_id = row[col["追踪编号"]]
        warehouse_raw = row[col["仓库"]]
        qty = row[col["数量"]] or 0
        results.append(
            ShippingPlanRow(
                sku=str(sku).strip(),
                warehouse_raw=str(warehouse_raw) if warehouse_raw is not None else "",
                warehouse_code=normalize_warehouse(warehouse_raw),
                tracking_id=str(tracking_id).strip() if tracking_id is not None else "",
                planned_quantity=int(qty),
            )
        )
    return results
