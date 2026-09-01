"""SKU 归一化 + 读取"翻译表"（WM-SKU → 原SKU/货号）。

翻译表目前是 测试8.24.xlsx，只用 "SKU" 和 "原sku" 两列——这份表本身还有箱数/已发货商品这些
统计列，那是之前人工核对用的，不是这次核对的输入，不读。
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

_TOKEN_PATTERN = re.compile(r"[A-Z]+|\d+")


def normalize_sku(value) -> str:
    """按"字母串 + 数字串"的顺序做归一化，用于匹配 key（不用于展示）。

    做法是把字母连续的部分和数字连续的部分依次提取出来，中间的分隔符——不管是 "-"、空格、
    还是 OCR 读错了变成的长破折号、双短横线之类——一律忽略，只要字母串和数字串本身的内容、
    数量、顺序完全一样就算同一个 SKU。比如 "WMTD-617"、"WMTD—-617"（OCR 常见的错读）、
    "wmtd 617" 都会归一化成一样的结果，不用针对每一种具体的错读方式单独打补丁。
    """
    if value is None:
        return ""
    text = str(value).upper()
    tokens = _TOKEN_PATTERN.findall(text)
    return "-".join(tokens)


def parse_translation(xlsx_path: str | Path, sheet_name: str | None = None) -> dict[str, str]:
    """返回 {归一化后的 WM-SKU: 原sku(货号，保留原始写法)}。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    try:
        sku_col = header.index("SKU")
        original_col = header.index("原sku")
    except ValueError as e:
        raise ValueError(f"翻译表缺少必要的列（SKU / 原sku），实际表头是：{header}") from e

    mapping: dict[str, str] = {}
    for values in rows_iter:
        wm_sku = normalize_sku(values[sku_col])
        rk_sku = values[original_col]
        if not wm_sku or rk_sku is None:
            continue
        mapping[wm_sku] = str(rk_sku).strip()
    return mapping
